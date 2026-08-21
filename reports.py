#####################################################################################################
#Funciones para la generación de informes en PDF, Excel y CSV, y archivo de eventos en formato .tsv #
#####################################################################################################

#imports genéricos
from dataclasses import dataclass, field

import customtkinter as ctk
import datetime
import statistics

from widgets import Canal, Consola
import config

#Imports necesarios para la generación del informe:
#se utilizan para generar el pdf:
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.enums import TA_CENTER

import openpyxl #se utiliza para generar un excel
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt
import tempfile
import os
import csv #ya se encuentraya incluido en python
from typing import Any
import matplotlib as mpl #para poder exportar las gráficas como imágenes


#------------------------------------------------------------
#Contenedor de datos para la generación de informes
#-----------------------------------------------------------
#módulo estándar que sirve para crear clases destinadas principalmente a guardar datos

@dataclass
class DatosInforme:
    #Identificación de la sesión
    id_sesion: str
    id_paciente: str
    duracion_sesion: str
    tiempo_inicio_sesion: float

    #Parámetros del protocolo
    num_ciclos: str
    tiempo_exposicion: str
    tiempo_desensibilizacion: str
    intervalo_ciclos: str
    tiempo_calibrado: str

    #Datos de la sesión
    historial_sesion: list[dict] 
    metricas_calibracion: dict[int, dict]
    colores_canales: list[str]


def media(lista):
    # Acepta tanto listas/iterables como valores numéricos.
    try:
        if lista is None:
            return 0
        if isinstance(lista, (int, float)):
            return lista
        # Evitar calcular la media de strings
        if isinstance(lista, str):
            return 0
        # Intentar iterar y calcular la media
        seq = list(lista)
        return statistics.mean(seq) if seq else 0
    
    except TypeError:
        return -1 

def _calculo_max_min(lista):
    if not lista:
        return 0,0
    try:
        return max(lista), min(lista)
    except (TypeError, ValueError):
        return 0,0

def _conversor_list_muestras_calibracion(datos: DatosInforme):
    muestras_calibracion = []

    if not datos.metricas_calibracion:
        return []
    for num_canal, metrica in datos.metricas_calibracion.items():
        
        tiempo_inicial_calibracion = metrica.get("tiempo_inicio", datos.tiempo_inicio_sesion)
        lista_flujo = list(metrica.get("flujo", []) or [])
        lista_concentracion = list(metrica.get("concentracion", []) or [])
        lista_velocidad = list(metrica.get("velocidad", []) or [])
        lista_latencia = list(metrica.get("latencia", []) or [])

        #en caso de que las muestras de los parámetros de interés sean diferentes
        for i in range(max(len(lista_flujo),len(lista_concentracion),len(lista_latencia),len(lista_velocidad))):
            muestras_calibracion.append({
                "timestamp" : tiempo_inicial_calibracion + i*config.MUESTREO_CALIBRACION,
                "onset": round(i*config.MUESTREO_CALIBRACION, 3),
                "canal": num_canal,
                "olor": metrica.get("olor",""),
                "estado": "activo",
                "flujo": lista_flujo[i] if i< len(lista_flujo) else 0.0,
                "concentracion": lista_concentracion[i] if i< len(lista_concentracion) else 0.0,
                "velocidad_motor": lista_velocidad[i] if i< len(lista_velocidad) else 0.0,
                "latencia": lista_latencia[i] if i< len(lista_latencia) else 0,
            })

    return muestras_calibracion



##############################
# CSV y _eventos.tsv
#############################
def generar_csv(ruta: str , datos: DatosInforme):
        seccion_metadatos=[["OlfaMetric - Informe de sesión"],
                         ["ID Sesión", datos.id_sesion],
                         ["ID Paciente", datos.id_paciente],
                         ["Fecha",datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                         ["Duración sesión", datos.duracion_sesion.replace("Duración de sesión: ","")],
                         [],
                         []]#fila vacía como separador
        
        
        seccion_protocolo=[["PARÁMETROS DEL PROTOCOLO"],
                            ["Número de ciclos", datos.num_ciclos],
                            ["Tiempo de exposición", datos.tiempo_exposicion],
                            ["Tiempo de desensibilización", datos.tiempo_desensibilizacion],
                            ["Tiempo de intervalo entre ciclos", datos.intervalo_ciclos],
                            [],
                            []
                            ]

        #RESULTADOS CALIBRACIÓN
        seccion_calibrado_resultados = [["RESULTADOS CALIBRACIÓN"],["Tiempo calibración (s)","Canal", "Olor", "Concentración (µg/m\u00B3)","Concentración máxima (µg/m\u00B3)", 
                                   "Concentración mínima (µg/m\u00B3)", 
                                   "Flujo (ml/min)","Velocidad (rpm)","Latencia (ms)"]]

        #si la variable metricas_calibracion existe ejecuta el siguiente código
        if datos.metricas_calibracion:
            for num_canal, metrica in datos.metricas_calibracion.items():
                seccion_calibrado_resultados.append([datos.tiempo_calibrado,
                                        datos.colores_canales[num_canal],
                                        metrica.get("olor",""),
                                        round(media(metrica.get("concentracion", 0)), 2),
                                        round(_calculo_max_min(metrica.get("concentracion", 0))[0],2),
                                        round(_calculo_max_min(metrica.get("concentracion", 0))[1],2),
                                          round(media(metrica.get("flujo", 0)), 2),
                                          round(media(metrica.get("velocidad",0)), 1),
                                          round(media(metrica.get("latencia",0)), 1)])
                
        seccion_calibrado_resultados.append([])
        seccion_calibrado_resultados.append([])

        #HISTORIAL DE CALIBRADO
        seccion_calibrado_historial = [["HISTORIAL DE CALIBRACIÓN"],["Hora","Onset (s)","Canal", "Olor","Concentración (µg/m\u00B3)", 
                                   "Flujo (ml/min)","Velocidad (rpm)","Latencia (ms)"]]
        for muestra in _conversor_list_muestras_calibracion(datos):
                seccion_calibrado_historial.append([
                    datetime.datetime.fromtimestamp(muestra["timestamp"]).strftime("%H:%M:%S"),
                    muestra["onset"],
                    datos.colores_canales[muestra["canal"]],
                    muestra.get("olor",""),  
                    round(muestra.get("concentracion", 0),2),              
                    round(muestra.get("flujo", 0),2),
                    round(muestra.get("velocidad_motor",0),1),
                    round(muestra.get("latencia",0),1)
                    ])

        seccion_calibrado_historial.append([])
        seccion_calibrado_historial.append([])


        #HISTORIAL DE PROTOCOLO
        seccion_protocolo_historial = [["HISTORIAL DEL PROTOCOLO"],["Onset (s)","Hora","Canal", "Olor",
                                   "Estado", "Velocidad (rpm)","Latencia (ms)"]]
        
        for metrica in datos.historial_sesion :
            #las claves definidas deben coincidir con las del simulador
            onset = round(metrica["timestamp"] - datos.tiempo_inicio_sesion,3)
            seccion_protocolo_historial.append([
                onset,
                datetime.datetime.fromtimestamp(metrica["timestamp"]).strftime("%H:%M:%S"),
                datos.colores_canales[metrica["canal"]],
                metrica.get("olor",""),
                metrica.get("estado",""),
                metrica.get("velocidad_motor",0),
                metrica.get("latencia",0)
            ])
            
        ruta_eventos = ruta.replace(".csv", "_eventos.tsv")
        _generar_csv_eventos(ruta_eventos, datos)
        
        #Se añade el BOM de UTF-8 al principio del archivo. Sin esto Excel en Windows 
        # no muestra correctamente los caracteres especiales como µ o ó.
        with open(ruta, "w", newline="", encoding="utf-8-sig") as fila:
            writer = csv.writer(fila, delimiter=";")
                
            writer.writerows(seccion_metadatos)
            writer.writerows(seccion_protocolo)
            writer.writerows(seccion_calibrado_resultados)
            writer.writerows(seccion_calibrado_historial)
            writer.writerows(seccion_protocolo_historial)



def _generar_csv_eventos(ruta: str, datos: DatosInforme):
        """
        CSV de eventos de estimulación compatible con neuroimagen.
        Columnas: onset, duration, trial_type
        Formato compatible con BIDS (Brain Imaging Data Structure),
        estándar creciente en neuroimagen.
        """
        eventos = [["onset","duration","trial_type","channel","odor"]]

        canal_inicio = {}

        for metrica in datos.historial_sesion:
            canal = metrica.get("canal", -1)
            estado = metrica.get("estado", "")
            olor = metrica.get("olor", "")
            timestamp = metrica.get ("timestamp", 0)
            onset = round(timestamp - datos.tiempo_inicio_sesion,3)
            #print(canal_inicio)
            #print(estado)
            if estado == "activo" and canal not in canal_inicio:
                #se registra el inicio de la estimulación
                canal_inicio[canal]= {
                    "onset": onset,
                    "timestamp": timestamp,
                    "odor": olor
                }
            elif estado == "inactivo" and canal in canal_inicio:
                inicio = canal_inicio.pop(canal)
                duracion = round(onset - inicio["onset"],3)
                eventos.append([
                    inicio["onset"],
                    duracion,
                    "estimulación" if canal != 2 else "desensibilización",
                    datos.colores_canales[canal],
                    inicio["odor"]
                ])

        with open(ruta, "w", newline="", encoding="utf-8-sig") as fila:
            writer = csv.writer(fila, delimiter="\t")    
            writer.writerows(eventos)


def generar_excel(ruta: str, datos: DatosInforme):
        
        workbook = openpyxl.Workbook()

        #Estilos
        estilo_titulo = Font(bold=True, size=14, color='FF01BDCE')
        estilo_cabecera = Font(bold=True, color='FFFFFFFF')
        relleno_cabecera = PatternFill('solid', fgColor='FF1E1E1E')
        #relleno_fila_par = PatternFill('solid', fgColor='FF2B2B2B')
        centrado = Alignment(horizontal='center', vertical='center')

        def aplicar_estilo_cabecera(celda,texto):
            celda.value = texto
            celda.font = estilo_cabecera
            celda.fill = relleno_cabecera
            celda.alignment = centrado

        def ajustar_ancho_columnas(hoja):
            for columna in hoja.columns:
                max_ancho = max(len(str(celda.value or "")) for celda in columna)
                hoja.column_dimensions[columna[0].column_letter].widht= max_ancho + 4

        ##HOJA 1: RESUMEN DE SESIÓN-----------------------------------------
        hoja_resumen = workbook.active
        hoja_resumen.title = "Resumen de sesión"

        #Datos de sesión
        hoja_resumen['A1'].value = "OlfaMetric - Informe de sesión"
        hoja_resumen['A1'].font = Font(bold=True, size=20, color='FF01BDCE')
        
        datos_sesion = [
            ("ID sesión", datos.id_sesion),
            ("ID paciente", datos.id_paciente),
            ("Fecha", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Duración de sesión", datos.duracion_sesion.replace("Duración de sesión: ",""))
        ]
        for contador, (clave,valor) in enumerate(datos_sesion, start=2):
            hoja_resumen[f'A{contador}'] = clave
            hoja_resumen[f'A{contador}'].font = Font(bold = True)
            hoja_resumen[f'B{contador}'] = valor
        
        #Parámetros del protocolo
        fila = 7
        hoja_resumen[f'A{fila}'] = "PARÁMETROS DEL PROTOCOLO"
        hoja_resumen[f'A{fila}'].font = estilo_titulo
        fila += 1

        datos_protocolo = [
            ("Número de ciclos", datos.num_ciclos),
            ("Tiempo de exposición", f'{datos.tiempo_exposicion} s'),
            ("Tiempo de desensibilización", f'{datos.tiempo_desensibilizacion} s'),
            ("Tiempo de intervalo entre ciclos", f'{datos.intervalo_ciclos} s')
        ]

        for clave, valor in datos_protocolo:
            hoja_resumen[f'A{fila}'] = clave
            hoja_resumen[f'A{fila}'].font = Font(bold=True)
            hoja_resumen[f'B{fila}'] = valor
            fila += 1
        
        #Resultados de calibración
        fila += 1
        hoja_resumen[f'A{fila}'] = "RESULTADOS CALIBRACIÓN"
        hoja_resumen[f'A{fila}'].font = estilo_titulo
        fila +=1

        #se definen las cabeceras comunes a todos los datos
        cabeceras_calibracion = ["Tiempo calibración (s)","Canal", "Olor", "Concentración (µg/m\u00B3)","Concentración máxima (µg/m\u00B3)", 
                                "Concentración mínima (µg/m\u00B3)", "Flujo medio (ml/min)",
                                "Velocidad media(rpm)","Latencia media(ms)"]
        
        #se le aplica el estilo creado previamente correspondiente a las cabeceras
        for columna, cabecera in enumerate(cabeceras_calibracion, start=1):
            aplicar_estilo_cabecera(hoja_resumen.cell(fila, columna),cabecera)

        #salto de fila
        fila +=1
        

        #en caso de existir los datos relativos a la calibración se introducen en sus correspondientes
        #celdas mediante un bucle anidado, por el que con cada fila par se aplica un relleno característico
        if datos.metricas_calibracion:
            for num_canal, metrica in datos.metricas_calibracion.items():
                fila_datos = [
                    datos.tiempo_calibrado,
                    datos.colores_canales[num_canal],
                    metrica.get("olor",""),
                    round(media(metrica.get("concentracion", 0)), 2),
                    round(_calculo_max_min(metrica.get("concentracion", 0))[0],2),
                    round(_calculo_max_min(metrica.get("concentracion", 0))[1],2),
                    round(media(metrica.get("flujo", 0)),2),
                    round(media(metrica.get("velocidad",0)),1),
                    round(media(metrica.get("latencia",0)),1)
                    ]
                for columna, valor in enumerate(fila_datos, start=1):
                    hoja_resumen.cell(fila, columna).value = valor
                    #if fila % 2 == 0:
                        #hoja_resumen.cell(fila,columna).fill = relleno_fila_par
                
                fila +=1

        ajustar_ancho_columnas(hoja_resumen)


        ##HOJA 2: HISTORIAL DE CALIBRACIÓN--------------------------------
        hoja_historial_calibrado = workbook.create_sheet("Historial Calibración")

        cabeceras_historial_calibracion = ["Hora", "Onset(s)","Canal", "Olor", "Estado","Concentración (µg/m\u00B3)", "Flujo (ml/min)",
                                     "Velocidad (rpm)", "Latencia (ms)"]
        
        for columna, cabecera in enumerate(cabeceras_historial_calibracion,start=1):
            aplicar_estilo_cabecera(hoja_historial_calibrado.cell(1,columna),cabecera)
            
        for fila, muestra in enumerate(_conversor_list_muestras_calibracion(datos),start = 2):
            fila_datos = [
            datetime.datetime.fromtimestamp(muestra["timestamp"]).strftime("%H:%M:%S"),
            muestra["onset"],  
            datos.colores_canales[muestra["canal"]],
            muestra["olor"],
            muestra["estado"],
            round(muestra["concentracion"],2),
            round(muestra["flujo"], 2),
            round(muestra["velocidad_motor"], 1),
            round(muestra["latencia"], 1),
            ]
            for columna, valor in enumerate(fila_datos, start=1):
                hoja_historial_calibrado.cell(fila, columna).value = valor


        ajustar_ancho_columnas(hoja_historial_calibrado)

                

        ##HOJA 3: HISTORIAL DE PROTOCOLO-----------------------------
        hoja_historial_protocolo = workbook.create_sheet("Historial de protocolo")
        hoja_historial_protocolo.title = "Historial de protocolo"

        cabeceras_historial_protocolo = ["Onset","Hora","Canal", "Olor", "Estado",
                                     "Velocidad (rpm)", "Latencia (ms)"]
        
        for columna, cabecera in enumerate(cabeceras_historial_protocolo,start=1):
            aplicar_estilo_cabecera(hoja_historial_protocolo.cell(1,columna),cabecera)

        for fila, metrica in enumerate(datos.historial_sesion, start=2):
            onset = round(metrica["timestamp"]- datos.tiempo_inicio_sesion,3)
            fila_datos = [
                onset,
                datetime.datetime.fromtimestamp(metrica["timestamp"]).strftime("%H:%M:%S"),  
                datos.colores_canales[metrica["canal"]],
                metrica.get("olor",""),
                metrica.get("estado",""),
                round(media(metrica.get("velocidad_motor", 0)), 1),
                round(media(metrica.get("latencia", 0)), 1),
            ]

            for columna, valor in enumerate(fila_datos, start=1):
                hoja_historial_protocolo.cell(fila, columna).value = valor


        ajustar_ancho_columnas(hoja_historial_protocolo)

        ##HOJA 4: GRÁFICAS CALIBRACIÓN -----------------------------------
        hoja_graficas = workbook.create_sheet("Gráficas de Calibrado")
        hoja_graficas.title = "Gráficas de Calibrado"
        hoja_graficas['A1'].font = estilo_titulo

        archivos_temporales = []
        if datos.metricas_calibracion:
            fila = 3
            for num_canal, metricas in datos.metricas_calibracion.items():
                color_canal = datos.colores_canales[num_canal]
                #muestras = metrica.get("muestras", {})

                for parametro, valores in metricas.items():
                    if parametro == "olor" or parametro == "tiempo_inicio":
                        continue
                    if not valores:
                        continue
                    #se generar grafícas mediante el paquetede matplotlib, y se guardan como imágenes temporales
                    figura, eje = plt.subplots(figsize=(8,3), dpi=80)
                    tiempo = [i * config.MUESTREO_CALIBRACION for i in range(len(valores))]
                    eje.plot(tiempo,valores, color="#01bdce")
                    eje.set_title(f'Canal {color_canal}/{metricas.get("olor","")} - {parametro}', fontdict={'fontsize': 10, 'fontweight': 'bold'})
                    eje.set_xlabel('Tiempo (s)')
                    eje.set_ylabel(parametro)
                    eje.grid(True, linestyle='--', alpha=0.5)
                    eje.set_facecolor('#2b2b2b')
                    eje.patch.set_facecolor('#1e1e1e')
                    eje.tick_params(colors="white")
                    eje.title.set_color("white")
                    eje.xaxis.label.set_color('white')
                    eje.yaxis.label.set_color('white')

                    for spine in eje.spines.values():
                        spine.set_color('white')

                    #Se guarda temporalmente la figura/gráfica generada, para luego insertarla en el Excel
                    with tempfile.NamedTemporaryFile(suffix=".png", delete= False) as temporal:
                        ruta_figura = temporal.name
                    archivos_temporales.append(ruta_figura)

                    figura.savefig(ruta_figura, bbox_inches='tight', facecolor=figura.get_facecolor())
                    plt.close(figura)

                    #Ahora se inserta la imagen en la hoja Excel de la fila correspondiente.
                    imagen = openpyxl.drawing.image.Image(ruta_figura)
                    imagen.anchor = f'A{fila}'
                    hoja_graficas.add_image(imagen)
                    fila += 18

        #guardamos el workbook con el conjunto de las hojas y datos generados en la ruta seleccionada
        # por el usuario (siempre, haya o no datos de calibración)
        workbook.save(ruta)
        #Se eliminan los archivos temporales usados para las imágenes
        for archivo in archivos_temporales:
            if os.path.exists(archivo):
                os.unlink(archivo)

###########################
# PDF
##########################
def generar_pdf(ruta: str, datos: DatosInforme):

        #Estilos
        estilos = getSampleStyleSheet()
        estilo_titulo = ParagraphStyle('Titulo',
                                    fontSize=25,
                                    textColor=colors.HexColor('#01BDCE'),
                                    alignment=TA_CENTER,          
                                    spaceAfter=20,
                                    fontName='Helvetica-Bold')

        estilo_seccion = ParagraphStyle('Seccion',
                                    fontSize=16,
                                    textColor=colors.HexColor('#01BDCE'),
                                    spaceBefore=15,
                                    spaceAfter=10,
                                    fontName='Helvetica-Bold')

        estilo_tabla_cabecera = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E1E1E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#EEEEEE'), colors.white]),  
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('PADDING', (0, 0), (-1, -1), 8),
        ])

        historia = []

        #TÍTULO
        historia.append(Paragraph("OlfaMetric - Informe de sesion", estilo_titulo)) 
        historia.append(Spacer(1, 4*mm))

        #DATOS DE SESIÓN
        historia.append(Paragraph("Datos de sesion", estilo_seccion))
        datos_sesion = [
            ["ID sesion", datos.id_sesion],
            ["ID paciente", datos.id_paciente],
            ["Fecha", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Duracion de sesion", datos.duracion_sesion]
        ]
        tabla_sesion = Table(datos_sesion, colWidths=(100*mm, 80*mm))
        tabla_sesion.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('PADDING', (0, 0), (-1, -1), 8),
            ]))
        
        historia.append(tabla_sesion)
        historia.append(Spacer(1, 10*mm))

        #PARÁMETROS DEL PROTOCOLO
        historia.append(Paragraph("Parametros del protocolo", estilo_seccion))
        datos_protocolo = [
            ["Numero de ciclos", datos.num_ciclos],
            ["Tiempo de exposicion", f'{datos.tiempo_exposicion} s'],
            ["Tiempo de desensibilizacion", f'{datos.tiempo_desensibilizacion} s'],
            ["Tiempo de intervalo entre ciclos", f'{datos.intervalo_ciclos} s']
        ]
        tabla_protocolo = Table(datos_protocolo, colWidths=(100*mm, 80*mm))
        tabla_protocolo.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('PADDING', (0, 0), (-1, -1), 8),
            ]))
        historia.append(tabla_protocolo)
        historia.append(Spacer(1, 10*mm))

        #RESULTADOS DE CALIBRACIÓN
        historia.append(Paragraph("Resultados de calibracion", estilo_seccion))
        cabeceras_calibracion = [[                                    
        "Tiempo (s)","Canal", "Olor", "Conc. med(ug/m³)", "Conc.max(ug/m³)",
        "Conc.min(ug/m³)", "Flujo med(ml/min)", "Vel. med (rpm)", "Latencia med(ms)"
        ]]
        datos_calibracion = []
        if datos.metricas_calibracion:
            for num_canal, metrica in datos.metricas_calibracion.items():
                datos_calibracion.append([
                    datos.tiempo_calibrado,
                    datos.colores_canales[num_canal],
                    metrica.get("olor", ""),
                    round(media(metrica.get("concentracion", [])),2),
                    round(_calculo_max_min(metrica.get("concentracion", []))[0], 2),
                    round(_calculo_max_min(metrica.get("concentracion", []))[1], 2),
                    round(media(metrica.get("flujo", [])), 2),
                    round(media(metrica.get("velocidad", [])), 1),
                    round(media(metrica.get("latencia", [])), 1),
                    ])
        tabla_calibracion = Table(cabeceras_calibracion + datos_calibracion,
                               colWidths=[20*mm,22*mm, 20*mm, 20*mm, 20*mm, 24*mm, 22*mm, 22*mm, 20*mm])
        tabla_calibracion.setStyle(estilo_tabla_cabecera)
        historia.append(tabla_calibracion)
        historia.append(PageBreak())

        #HISTORIAL DE CALIBRACIÓN
        historia.append(Paragraph("Historial Calibración", estilo_seccion))
        cabeceras_historial_calibrado = [[
            "Hora", "Onset (s)","Canal", "Olor","Estado","Concentración (µg/m\u00B3)", "Flujo (ml/min)","Velocidad(rpm)", "Latencia(ms)"
        ]]
        datos_historial_calibrado = []
        for muestra in _conversor_list_muestras_calibracion(datos):
            datos_historial_calibrado.append([
                datetime.datetime.fromtimestamp(muestra["timestamp"]).strftime("%H:%M:%S"),
                muestra["onset"],
                datos.colores_canales[muestra["canal"]],
                muestra["olor"],
                muestra["estado"],
                round(muestra["concentracion"], 2),
                round(muestra["flujo"], 2),
                round(muestra["velocidad_motor"], 1),
                round(muestra["latencia"], 1),
                ])
        tabla_historial_calibracion = Table(cabeceras_historial_calibrado + datos_historial_calibrado, 
                                          colWidths=[16*mm, 20*mm, 22*mm, 24*mm, 20*mm, 24*mm, 20*mm, 20*mm, 20*mm])
        
        tabla_historial_calibracion.setStyle(estilo_tabla_cabecera)
        historia.append(tabla_historial_calibracion)
        historia.append(PageBreak())


        #HISTORIAL DE PROTOCOLO
        historia.append(Paragraph("Historial de datos", estilo_seccion))
        cabeceras_historial_protocolo = [[
            "Onset (s)","Hora", "Canal", "Olor","Estado","Velocidad(rpm)", "Latencia(ms)"
        ]]
        datos_historial_protocolo = []
        for metrica in datos.historial_sesion:
            onset = round(metrica["timestamp"] - datos.tiempo_inicio_sesion,3)
            datos_historial_protocolo.append([
                onset,
                datetime.datetime.fromtimestamp(metrica["timestamp"]).strftime("%H:%M:%S"),
                datos.colores_canales[metrica["canal"]],
                metrica.get("olor", ""),
                metrica.get("estado", ""),
                round(metrica.get("velocidad_motor", 0), 1),
                round(metrica.get("latencia", 0), 1),
                ])
        tabla_historial_protocolo = Table(cabeceras_historial_protocolo + datos_historial_protocolo, 
                                          colWidths=[20*mm, 20*mm, 22*mm, 22*mm, 20*mm, 22*mm, 20*mm])
        
        tabla_historial_protocolo.setStyle(estilo_tabla_cabecera)
        historia.append(tabla_historial_protocolo)

        #GENERAR PDF
        documento = SimpleDocTemplate(ruta, pagesize=A4,
                                       rightMargin=15*mm, leftMargin=15*mm,
                                       topMargin=15*mm, bottomMargin=15*mm)
        documento.build(historia)
            


