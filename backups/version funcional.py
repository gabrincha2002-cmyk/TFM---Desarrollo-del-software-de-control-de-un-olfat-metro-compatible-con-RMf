"""
=============================================================
 App de escritorio — Control remoto 6 motores 
 Controlador: ESP32 ESP-WROOM-32
 Dependencias: pip install customtkinter 
=============================================================
"""

import customtkinter as ctk
import time

import datetime
import math

#Imports necesarios para la generación del informe:
#se utilizan para generar el pdf:
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.enums import TA_CENTER

import openpyxl #se utiliza para generar un excel
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt
import tempfile
import os
import csv #ya se encuentraya incluido en python
import matplotlib as mpl #para poder exportar las gráficas como imágenes


#para gráficas en tiempo real
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import collections

#conexión con simulación del ESP32
from ws_client_original import WSClient

#para generar un hilo paralelo a la app
import threading
import zeroconf as zc


#para el cálculo de métricas de calibración
import statistics

#para generar números aleatorios en el protocolo aleatorio
import random





ctk.set_appearance_mode("dark")  # Modos: "System" (predeterminado), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Temas: "blue" (predeterminado), "dark-blue", "green"

# ─────────────────────────────────────────────────────────────
#  FUNCIONALIDADES
# ─────────────────────────────────────────────────────────────



# ─────────────────────────────────────────────────────────────
#  INTERFAZ GRÁFICA
# ─────────────────────────────────────────────────────────────
class Canal(ctk.CTkFrame):
    def __init__(self, master, color_canal, num_canal, registro=None, actualizar_canal=None):
        super().__init__(master, fg_color="#343638", border_color="#4a4c4e", border_width=1, corner_radius=10)
        self.color_canal = color_canal
        self.num_canal = num_canal
        self.actualizar_canal = actualizar_canal
        self.registro = registro
        self.after_cronometro = None
        self.estado_canal= False
        self.crear_canal()

    def crear_canal(self):
        self.grid_columnconfigure((0,1), weight=1)
        self.grid_rowconfigure((0,1), weight=1)
        
        # Color canal
        self.l_color_canal = ctk.CTkLabel(self, text=f"Canal {self.color_canal}", font=ctk.CTkFont(size=20, weight="bold"))
        self.l_color_canal.grid(row=0, column=0, padx=(10,5), pady=10, sticky="w")

        # Olor del canal
        self.e_olor_canal = ctk.CTkEntry(self, placeholder_text="Olor del canal", font=ctk.CTkFont(size=20))
        self.e_olor_canal.grid(row=0, column=0, padx=10, pady=10, sticky="e")

        #Tiempo activo del canal
        self.sv_tiempo_activo=ctk.StringVar(value="Tiempo activo: 00:00:00")  # Variable para almacenar el tiempo activo del canal
        self.l_tiempo_act_canal = ctk.CTkLabel(self, textvariable= self.sv_tiempo_activo, font=ctk.CTkFont(size=14))
        self.l_tiempo_act_canal.grid(row=2, column=0, padx=10, pady=(10,5), sticky="w")

        # Barra de progreso de la actividad
        self.pb_actividad_canal = ctk.CTkProgressBar(self, width=400, height=20)
        self.pb_actividad_canal.grid(row=3, column=0, padx=10, pady=(10,5), sticky="w")

        # Porcentaje de actividad
        self.l_porcentaje_act_canal = ctk.CTkLabel(self, text="0%", font=ctk.CTkFont(size=14))
        self.l_porcentaje_act_canal.grid(row=3, column=1, padx=10, pady=(5,10), sticky="w")

        #Botón Activar

        self.b_activar_canal = ctk.CTkButton(self, text="Activar", fg_color="#85ad75",text_color="#ffffff", 
                                             hover_color="#488f51",corner_radius=10,border_color="#006400",border_width=1,
                                             font=ctk.CTkFont(size=16, weight="bold"), command=self.activar_canal)
        self.b_activar_canal.grid(row=5, column=0, padx=10, pady=(5,10), sticky="w")

        self.b_parar_canal = ctk.CTkButton(self, text="Parar", fg_color="#f56a6a",text_color="#ffffff",
                                             hover_color="#ee4242",corner_radius=10,border_color="#ff0000",border_width=1,
                                             font=ctk.CTkFont(size=16, weight="bold"), command=self.parar_canal)
        self.b_parar_canal.grid(row=5, column=1, padx=10, pady=(5,10), sticky="w")

    def cronometro(self,tiempo):
        if self.estado_canal:
            self.sv_tiempo_activo.set(f"Tiempo activo: {tiempo.strftime('%H:%M:%S')}")
            self.after_cronometro = self.after(1000, lambda: self.cronometro(tiempo + datetime.timedelta(seconds=1)))

    def resetear_cronometro(self):
        self.sv_tiempo_activo.set("Tiempo activo: 00:00:00")

    def pausar_cronometro(self):
        if self.after_cronometro:
            self.after_cancel(self.after_cronometro)
            self.after_cronometro = None

    def activar_canal(self,tiempo_inicial=datetime.datetime.strptime("00:00:00", "%H:%M:%S")):
        if not self.estado_canal:
            self.estado_canal=True
            self.cronometro(tiempo_inicial)
            self.configure(fg_color="#256F2F", border_color="#7DEB7D", border_width=4)  # Cambia el fondo del canal para indicar que está activo
            self.b_activar_canal.configure(text="En marcha", fg_color="#70c64e",text_color="#ffffff",border_color="#006400",border_width=1,font=ctk.CTkFont(size=14, weight="bold"))
            #FALTA INCLUIR FUNCIONALIDAD PARA ACTIVAR EL CANAL EN EL ESP32
            if self.registro:
                self.registro(f"Canal {self.color_canal} ({self.e_olor_canal.get()}) ACTIVADO")
            
            if self.actualizar_canal:
                self.actualizar_canal(self.num_canal,"activar")
            

    def parar_canal(self):
        if self.estado_canal:
            self.estado_canal=False
            self.pausar_cronometro()
            self.configure(fg_color="#343638", border_color="#4a4c4e", border_width=1)  # Restaura el fondo del canal para indicar que está inactivo
            self.b_activar_canal.configure(text="Activar", fg_color="#85ad75",text_color="#ffffff",border_color="#006400",border_width=1,
                font=ctk.CTkFont(size=14, weight="bold"))
            #FALTA INCLUIR FUNCIONALIDAD PARA PARAR EL CANAL EN EL ESP32
            #self.canal_anterior.set(f"Canal {self.color_canal} ({self.e_olor_canal.get()})")
            #print(f"Canal {self.color_canal} ({self.e_olor_canal.get()}")
            if self.registro:
                self.registro(f"Canal {self.color_canal} ({self.e_olor_canal.get()}) DETENIDO")
            
            if self.actualizar_canal:
                self.actualizar_canal(self.num_canal,"parar")
            


class SpinboxCTk(ctk.CTkFrame):
    def __init__(self, master, valor=120,valor_min=0, valor_max=9999999, escalon=1
                 , border_width=0, corner_radius=10, width=120, height=30):
        super().__init__(master, fg_color = "#242424",width=width, height=height, border_width=border_width, corner_radius=corner_radius)
        self.valor_min = valor_min
        self.valor_max = valor_max
        self.escalon = escalon
        self.valor = ctk.StringVar(value=valor)
        self.crear_spinboxCTk()

    def crear_spinboxCTk(self):
        f_spinbox = ctk.CTkFrame(self,fg_color="#242424", bg_color="#242424")
        f_spinbox.grid_columnconfigure((0,2),weight=0)
        f_spinbox.grid_columnconfigure(1,weight=0)
        f_spinbox.grid(row=0, column=0, sticky="w", padx=10, pady=10,)

        self.b_decrementar = ctk.CTkButton(f_spinbox,bg_color="#242424",fg_color="#0a5f70", text="-", width=30, command=self.decrementar,font=ctk.CTkFont(size=14, weight="bold"),border_color="#0a5f70", border_width=1)
        self.b_decrementar.grid(row=0,column=0, padx=5, pady=5, sticky="w")

        self.e_spinbox = ctk.CTkEntry(f_spinbox,bg_color="#242424", fg_color="#242424", textvariable=self.valor, font=ctk.CTkFont(size=14))
        self.e_spinbox.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        self.b_incrementar = ctk.CTkButton(f_spinbox, bg_color="#242424", fg_color="#0a5f70",text="+", width=30, command=self.incrementar,font=ctk.CTkFont(size=14, weight="bold"),border_color="#0a5f70", border_width=1)
        self.b_incrementar.grid(row=0,column=2, padx=5, pady=5, sticky="e")

    def incrementar(self):
        try:
            valor_actual = int(self.valor.get())
            valor_final = valor_actual + self.escalon
            if valor_final <= self.valor_max:
                self.valor.set(valor_final)
        except ValueError:
            pass

    def decrementar(self):
        try:
            valor_actual = int(self.valor.get())
            valor_final = valor_actual - self.escalon
            if valor_final >= self.valor_min:
                self.valor.set(valor_final)
        except ValueError:
            pass
    
    def get(self):
        try:
            return int(self.valor.get())
        except ValueError:
            return self.valor_min
        
    def set(self, valor):
        self.valor.set(int(valor)) 

class FrameDeslizante(ctk.CTkScrollableFrame):
    def __init__(self, master,fg_color,border_color="#4a4c4e", border_width=1, corner_radius=10):
        super().__init__(master,fg_color=fg_color,border_color=border_color, border_width=border_width, corner_radius=corner_radius)
        self.fg_color = fg_color
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)  # Cabecera se expande
        
class Consola(ctk.CTkTextbox):
    def __init__(self, master):
        super().__init__(master)
        self.crear()

    def crear(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(self, text="Consola", text_color="#458B8D", font=ctk.CTkFont(size=22, weight="bold")).grid(row=0, column=0, padx=5, pady=(5,5), sticky="nw")

        self.t_registro=ctk.CTkTextbox(self, width=1000, height=150, font=ctk.CTkFont(size=14),)
        self.t_registro.grid(row=1, column=0, padx=10, pady=(10,5), sticky="nsew")

    def registro(self, mensaje, nivel="INFO"):
        hora = datetime.datetime.now().strftime("%H:%M:%S")
        mensaje_final = f"{hora}- {nivel}- {mensaje}\n"
        self.t_registro.insert(index="end", text=mensaje_final)
        self.t_registro.see("end")

    def limpiar_registro(self):
        self.t_registro.delete("1.0","end")  
        

    
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("OlfaMetric")
        self.geometry("1920x1080")
        self.colores_canales = ["Verde Claro", "Negro", "Blanco", "Azul", "Amarillo", "Rojo"]  # Colores para cada cartucho
        self.tiempo_grafica_flujo = list(range(61))
        self.tiempo_grafica_concentracion = list(range(61))
        self.tiempo_grafica_latencia = list(range(61))
        self.tiempo_grafica_velocidad = list(range(61))

        #buffers termporales
        self.olores = []
        self.historial_sesion= []
        self.metricas_calibracion = {}
        # Los widgets de UI (p. ej. `e_tiempo_calibrado`) se crean en `crear_ui()` más abajo,
        # por eso no debemos usar `self.e_tiempo_calibrado.get()` aquí (aún no existe).
        # Usar el tamaño por defecto de las gráficas (`self.tiempo_grafica_flujo`) para inicializar buffers.
        self.buffer_flujo = collections.deque([0]*61, maxlen=61)
        self.buffer_concentracion = collections.deque([0]*61,maxlen=61)
        self.buffer_latencia = collections.deque([0]*61,maxlen=61)
        self.buffer_velocidad = collections.deque([0]*61,maxlen=61)

    
        #buffers históricos
        self.buffer_historico_flujo = []
        self.buffer_historico_concentracion = []
        self.buffer_historico_velocidad = []
        self.buffer_historico_latencia = []
        self.buffer_historico_timestamps = []
        self.buffer_historico_olores = []

        self.segundos_restantes_protocolo = 0
        self.segundos_restantes_velocidad = 0
        self.segundos_restantes_flujo = 0
        self.segundos_restantes_concentracion = 0
        self.segundos_restantes_latencia = 0
        self.tiempo_guardado = None

        #protocolo
        self.after_activo = None
        self.canal_activo= None
        self.canal_anterior= None
        self.canal_siguiente= None
        self.sv_canal_activo = ctk.StringVar(value="Ninguno")
        self.sv_canal_anterior = ctk.StringVar(value="Ninguno")
        self.sv_canal_siguiente = ctk.StringVar(value="Ninguno")
        self.protocolo_activo = False
        self.protocolo_parado = False
        self.orden_protocolo = None

        #calibrado
        self.canal_calibrado = None

        self.calibrado_velocidad_parado = False
        self.calibrado_flujo_parado = False
        self.calibrado_concentracion_parado = False
        self.calibrado_latencia_parado = False

        # flags por parámetro
        self.calibrando_velocidad        = False
        self.calibrando_flujo            = False
        self.calibrando_concentracion    = False
        self.calibrando_latencia         = False
        # buffers por parámetro
        self.buffer_historico_calibrado_velocidad      = []
        self.buffer_historico_calibrado_flujo          = []
        self.buffer_historico_calibrado_concentracion  = []
        self.buffer_historico_calibrado_latencia       = []
        # after por parámetro
        self.after_calibrado_velocidad       = None
        self.after_calibrado_flujo           = None
        self.after_calibrado_concentracion   = None
        self.after_calibrado_latencia        = None
        self.after_actualizar_graficas       = None

        #Cliente para establecer conexión con el controlador ESP32-WROOM (o simulador)
        self.ws_client = WSClient(
        uri        = "ws://localhost:8765",   # o IP del ESP32
        on_datos   = self._on_datos_ws,
        on_estado  = self._on_estado_ws,
    )

        self.grid_rowconfigure(0, weight=0)  # Cabecera
        self.grid_rowconfigure(1, weight=1)  # Canales y Protocolo se expande
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=0)  # Consola no se expande

        self.grid_columnconfigure(0, weight=2)  # Columna de canales y cabecera se expande
        self.grid_columnconfigure(1, weight=1) # Columna de protocolo se expande 

        self.ws_client.iniciar()
        self.crear_ui()
        self.tiempo_sesion()




    #CONSTRUCCIÓN DE INTERFAZ DE USUARIO
    def crear_ui(self):
        #Para crear un marco dentro de la ventana principal
        self.f_cabecera = ctk.CTkFrame(self,fg_color="#1e1e1e",border_color="#4a4c4e", border_width=1, corner_radius=10)  # Fondo transparente para el marco
        self.f_cabecera.grid(row=0,column=0,sticky="nsew",columnspan=2)  # Usamos grid para colocar el marco

        self.f_cabecera.grid_columnconfigure((0,1), weight=1) 
        self.f_cabecera.grid_rowconfigure(0, weight=1)
        #self.f_cabecera.grid_rowconfigure(1, weight=1)
        #self.f_cabecera.grid_columnconfigure(2, weight=0) 


        #CABECERA
        self.l_titulo = ctk.CTkLabel(self.f_cabecera,text="OlfaMetric",corner_radius=10,
                                  text_color="#01bdce", font=ctk.CTkFont(size=60, overstrike=False, weight="bold"))
        self.l_titulo.grid(row=0, column=0, padx=20, pady=10, sticky="w")
        self.l_subtitulo = ctk.CTkLabel(self.f_cabecera,text="- An olfactory metric for everyone -",corner_radius=10,
                                  text_color="#80E8E9", font=ctk.CTkFont(size=30, weight="normal",slant="italic"))
        self.l_subtitulo.grid(row=0, column=0, padx=350, pady=5, sticky="w")

        self.l_estado_conexion = ctk.CTkLabel(self.f_cabecera,text="○ Desconectado",fg_color="#1e1e1e",text_color="#fa8989",
                                     font=ctk.CTkFont(size=18,weight="bold"))
        self.l_estado_conexion.grid(row=0, column=1,padx=30, pady=10, sticky="en") 

        self.b_buscar_dispositivos = ctk.CTkButton(self.f_cabecera, text="Buscar dispositivos", fg_color="#1e1e1e",
                                                 text_color="#828282",corner_radius=10,border_width=1,
                                                  command=self.b_buscar_dispositivos,font=ctk.CTkFont(size=14, weight="bold"))
        self.b_buscar_dispositivos.grid(row=0, column=1, padx=10, pady=(0,5), sticky="es")

        #CONSOLA
        #La creo antes que canales para que la creación de estos no me den ningún problema (están relacionados por
        # la función registro de la clase Consola) 
        self.f_consola = ctk.CTkFrame(self, fg_color="#1e1e1e",border_color="#4a4c4e", border_width=1, corner_radius=10, height=100)
        self.f_consola.grid(row=3,column=0,columnspan=1,padx=5,pady=(5,10),sticky="ew")
        self.consola= Consola(self.f_consola)
        self.consola.grid(row=0,column=0,padx=10,pady=(10,5),sticky="ew")


        
        self.tv_canales_calibracion = ctk.CTkTabview(master=self, fg_color="transparent",border_color="#4a4c4e", border_width=1, corner_radius=10, width=450)
        self.tv_canales_calibracion.grid(row=1,column=0,padx=5,pady=10,sticky="nsew", rowspan=2)
        self.tv_canales_calibracion.add("Canales")
        self.tv_canales_calibracion.add("Calibración")
        self.tv_canales_calibracion._segmented_button.configure(text_color="#ffffff", border_width=1, corner_radius=10, width = 200,height =30, font=ctk.CTkFont(size=20, weight="bold"))

        self.tv_canales_calibracion.tab("Canales").grid_columnconfigure(0, weight=1)
        self.tv_canales_calibracion.tab("Calibración").grid_columnconfigure(0, weight=1)

        self.tv_canales_calibracion.set("Canales")

        self.f_canales = ctk.CTkScrollableFrame(self.tv_canales_calibracion.tab("Canales"),fg_color="transparent")  # Fondo transparente para el marco
        self.f_canales.grid(row=0,column=0,padx=5,pady=10,sticky="nsew",rowspan=2)  # Usamos grid para colocar el marco
        self.tv_canales_calibracion.tab("Canales").grid_rowconfigure(0, weight=1)
        self.tv_canales_calibracion.tab("Canales").grid_columnconfigure(0, weight=1)
        self.f_canales.grid_columnconfigure(0, weight=1)
        self.f_canales.grid_columnconfigure(1, weight=1)




        #self.l_canales= ctk.CTkLabel(self.f_canales, text="Canales",text_color="#458B8D", font=ctk.CTkFont(size=22, weight="bold"))
        #self.l_canales.grid(row=0, column=0, padx=10, pady=(10,5), sticky="w")

         #CANALES
        #-----------
        self.cuadros_canales = []
        for i in range(0,6):
            columna = i//3
            fila = i%3
            cuadro_canal= Canal(self.f_canales,color_canal=self.colores_canales[i],num_canal=i,registro=self.consola.registro, actualizar_canal=self.actualizar_canales)
            cuadro_canal.grid(row=fila+1, column=columna, padx=10, pady=10, sticky="nsew")
            #actualiza el CombBox de calibración cada vez que se modifica el nombre del canal
            cuadro_canal.e_olor_canal.bind("<FocusOut>", lambda e: self.actualizar_comb_canales_calibrados())
            self.cuadros_canales.append(cuadro_canal)
       
        self.cuadros_canales[self.colores_canales.index("Blanco")].e_olor_canal.destroy()  # El canal blanco no tiene entrada de olor, por lo que se destruye el widget de entrada

        #----------
        #CALIBRACIÓN
        # Crea un frame deslizante dentro del tabview
        self.f_calibracion_scroll = ctk.CTkScrollableFrame(self.tv_canales_calibracion.tab("Calibración"), fg_color="transparent")
        self.f_calibracion_scroll.grid(row=0, column=0, padx=0, pady=0, sticky="nsew")
        self.tv_canales_calibracion.tab("Calibración").grid_rowconfigure(0, weight=1)
        self.tv_canales_calibracion.tab("Calibración").grid_columnconfigure(0, weight=1)
        self.f_calibracion_scroll.grid_columnconfigure(0, weight=1)
        self.f_calibracion_scroll.grid_columnconfigure(1, weight=1)
        #self.f_calibracion_scroll.grid_rowconfigure(2, weight=1)
        #self.f_calibracion_scroll.grid_rowconfigure(3, weight=1)   

        
        #self.l_calibracion = ctk.CTkLabel(self.f_calibracion_scroll, text="Calibración", text_color="#458B8D", font=ctk.CTkFont(size=22, weight="bold"))
        #self.l_calibracion.grid(row=0, column=0, padx=10, pady=(10,5), sticky="n")

        # Selección de canales a calibrar
        sv_canales_calibrado = ctk.StringVar(value="Seleccionar canal a calibrar")
        self.comb_canales_calibrados = ctk.CTkComboBox(self.f_calibracion_scroll, values=[],
                                             width=180, height=36, text_color="#ffffff",command=self.seleccionar_calibrado,dropdown_fg_color="#0a5f70", variable=sv_canales_calibrado)
        self.comb_canales_calibrados.grid(row=0, column=0, padx=0, pady=5, sticky="w")

        self.e_tiempo_calibrado = SpinboxCTk(self.f_calibracion_scroll, valor=30, valor_min=1, valor_max=9999999, escalon=1, width=60, height=15)
        self.e_tiempo_calibrado.grid(row=0, column=0, padx=0, pady=5, sticky="e")
        self.l_tiempo_calibrado = ctk.CTkLabel(self.f_calibracion_scroll, text="Tiempo:", text_color="#458B8D", font=ctk.CTkFont(size=14, weight="bold"))
        self.l_tiempo_calibrado.grid(row=0, column=0, padx=195, pady=5, sticky="w")



        #Iniciar calibrado general
        self.b_iniciar_calibrado_general = ctk.CTkButton(self.f_calibracion_scroll, text="Inicio General", text_color="#ffffff", fg_color= "#85ad75", width= 10, height=15,
                                                   hover_color="#488f51", border_color="#006400", border_width=1, corner_radius=10,command=self.iniciar_calibrado_general,font=ctk.CTkFont(size=18, weight="bold"))
        self.b_iniciar_calibrado_general.grid(row=0, column=1, padx=(24,0), pady=5, sticky="w")

        #reiniciar calibrado general
        self.b_reiniciar_calibrado_general = ctk.CTkButton(self.f_calibracion_scroll, text="Reinicio General", text_color="#ffffff", fg_color= "#C7BE19" , width=10, height=15,
                                                   hover_color="#9da31e", border_color="#F6F04F", border_width= 1, corner_radius=10,command=self.reiniciar_calibrado_general,font=ctk.CTkFont(size=18, weight="bold"))
        self.b_reiniciar_calibrado_general.grid(row=0, column=1, padx=(160,0), pady=5, sticky="w")

        #parar calibrado general
        self.b_parar_calibrado_general = ctk.CTkButton(self.f_calibracion_scroll, text="Parada General", text_color="#ffffff", fg_color= "#f56a6a", width=10, height=15,
                                                   hover_color="#ee4242", border_color="#ff0000", border_width= 1, corner_radius=10,command=self.parar_calibrado_general,font=ctk.CTkFont(size=18, weight="bold"))
        self.b_parar_calibrado_general.grid(row=0, column=1, padx=(0,30), pady=5, sticky="e")

        #Calibración Velocidad
        self.l_calibrado_velocidad = ctk.CTkLabel(self.f_calibracion_scroll, text="Velocidad", text_color="#458B8D", font=ctk.CTkFont(size=20, weight="bold"))
        self.l_calibrado_velocidad.grid(row=1, column=0, padx=60, pady=0, sticky="ws")

        #Gráfica en tiempo real - Velocidad
        fig_velocidad = Figure(figsize=(10,5),dpi=60)
        self.ax_velocidad = fig_velocidad.add_subplot(111)
        self.ax_velocidad.set_xlabel("Tiempo (s)")
        self.ax_velocidad.set_ylabel("Velocidad (m/s)")
        #fig_velocidad.set_facecolor("#222526")  # Fondo de la figura transparente
        #self.ax_velocidad.set_facecolor("#222526")  # Fondo del gráfico transparente
        fig_velocidad.set_facecolor("#242424")  # Fondo de la figura transparente
        self.ax_velocidad.set_facecolor("#242424")
        self.ax_velocidad.tick_params(colors = "#ffffff") 
        self.ax_velocidad.xaxis.label.set_color("#ffffff")  # Color de las etiquetas del eje x
        self.ax_velocidad.yaxis.label.set_color("#ffffff")  # Color de las etiquetas del eje y
        self.ax_velocidad.spines['bottom'].set_color('#ffffff')  # Color de la línea del eje x
        self.ax_velocidad.spines['left'].set_color('#ffffff')  # Color de
        self.ax_velocidad.spines['top'].set_color('#ffffff')  # Color de la línea superior
        self.ax_velocidad.spines['right'].set_color('#ffffff')  # Color de la línea derecha
        self.ax_velocidad.plot(self.tiempo_grafica_velocidad,self.buffer_velocidad)
        fig_velocidad.tight_layout()
        self.canvas_velocidad = FigureCanvasTkAgg(fig_velocidad, master=self.f_calibracion_scroll)
        self.canvas_velocidad.get_tk_widget().grid(row=2, column=0, padx=10, pady=0, sticky="nesw")
        self.canvas_velocidad.draw()

        self.f_botones_calibrado_velocidad = ctk.CTkFrame(self.f_calibracion_scroll, fg_color="transparent", width=75, height=15, corner_radius=10,border_width=1)
        self.f_botones_calibrado_velocidad.grid(row=1, column=0, padx=10, pady=0, sticky="se")
        self.f_botones_calibrado_velocidad.grid_columnconfigure(0, weight=1)
        self.f_botones_calibrado_velocidad.grid_columnconfigure(1, weight=1)
        self.f_botones_calibrado_velocidad.grid_columnconfigure(2, weight=1)


        self.b_iniciar_calibrado_velocidad = ctk.CTkButton(self.f_botones_calibrado_velocidad, text="▶", text_color = "#ffffff", fg_color="#85ad75", width=25, height=15, 
        hover_color="#488f51", border_color="#006400", border_width=1, corner_radius=5,command=self.iniciar_calibrado_velocidad,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_iniciar_calibrado_velocidad.grid(row=0, column=0, ipadx = 0, padx=0, pady=0, sticky="nswe")
       

        self.b_reiniciar_calibrado_velocidad = ctk.CTkButton(self.f_botones_calibrado_velocidad, text="↻" , text_color = "#ffffff", fg_color="#C7BE19",width=25, height=15,
                                                   hover_color="#9da31e", border_color="#CFC61B", border_width=1, corner_radius=5,command=self.reiniciar_calibrado_velocidad,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_reiniciar_calibrado_velocidad.grid(row=0, column=1, ipadx = 0, padx=0, pady=0, sticky="nswe")

        self.b_parar_calibrado_velocidad = ctk.CTkButton(self.f_botones_calibrado_velocidad, text="◼", text_color = "#ffffff",width=25, height=15, 
        fg_color="#f56a6a", hover_color = "#ee4242", border_color="#ff0000",corner_radius=5,border_width=1, command=self.parar_calibrado_velocidad,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_parar_calibrado_velocidad.grid(row=0, column=2, ipadx = 0 ,padx=0, pady=0, sticky="nswe")


        #Calibración Flujo
        self.l_calibrado_flujo = ctk.CTkLabel(self.f_calibracion_scroll, text="Flujo", text_color="#458B8D", font=ctk.CTkFont(size=20, weight="bold"))
        self.l_calibrado_flujo.grid(row=1, column=1, padx=60, pady=0, sticky="ws")

        #Gráfica en tiempo real - Flujo
        fig_flujo = Figure(figsize=(10,5),dpi=60)
        self.ax_flujo = fig_flujo.add_subplot(111)
        self.ax_flujo.set_xlabel("Tiempo (s)")
        self.ax_flujo.set_ylabel("Flujo (ml/min)")
        self.ax_flujo.plot(self.tiempo_grafica_flujo,self.buffer_flujo)
        fig_flujo.set_facecolor('#242424')  # Fondo de la figura transparente
        self.ax_flujo.set_facecolor("#242424")  # Fondo del gráfico transparente
        self.ax_flujo.tick_params(colors = "#ffffff") 
        self.ax_flujo.xaxis.label.set_color("#ffffff")  # Color de las etiquetas del eje x
        self.ax_flujo.yaxis.label.set_color("#ffffff")  # Color de las etiquetas del eje y
        self.ax_flujo.spines['bottom'].set_color('#ffffff')  # Color de la línea del eje x
        self.ax_flujo.spines['left'].set_color('#ffffff')  # Color de
        self.ax_flujo.spines['top'].set_color('#ffffff')  # Color de la línea superior
        self.ax_flujo.spines['right'].set_color('#ffffff')  # Color de la línea derecha
        fig_flujo.tight_layout()
        self.canvas_flujo = FigureCanvasTkAgg(fig_flujo, master=self.f_calibracion_scroll)
        self.canvas_flujo.get_tk_widget().grid(row=2, column=1, padx=10, pady=0, sticky="nsew")
        self.canvas_flujo.draw()

        self.f_botones_calibrado_flujo = ctk.CTkFrame(self.f_calibracion_scroll, fg_color="transparent", width=75, height=15, corner_radius=10,border_width=1)
        self.f_botones_calibrado_flujo.grid(row=1, column=1, padx=10, pady=0, sticky="se")
        self.f_botones_calibrado_flujo.grid_columnconfigure(0, weight=1)
        self.f_botones_calibrado_flujo.grid_columnconfigure(1, weight=1)
        self.f_botones_calibrado_flujo.grid_columnconfigure(2, weight=1)

        self.b_iniciar_calibrado_flujo = ctk.CTkButton(self.f_botones_calibrado_flujo, text="▶", text_color = "#ffffff", fg_color="#85ad75", width=25, height=15, 
        hover_color="#488f51", border_color="#006400", border_width=1, corner_radius=5,command=self.iniciar_calibrado_flujo,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_iniciar_calibrado_flujo.grid(row=0, column=0, ipadx = 0, padx=0, pady=0, sticky="nswe")
       

        self.b_reiniciar_calibrado_flujo = ctk.CTkButton(self.f_botones_calibrado_flujo, text="↻" , text_color = "#ffffff", fg_color="#C7BE19",width=25, height=15,
                                                   hover_color="#9da31e", border_color="#CFC61B", border_width=1, corner_radius=5,command=self.reiniciar_calibrado_flujo,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_reiniciar_calibrado_flujo.grid(row=0, column=1, ipadx = 0, padx=0, pady=0, sticky="nswe")

        self.b_parar_calibrado_flujo = ctk.CTkButton(self.f_botones_calibrado_flujo, text="◼", text_color = "#ffffff",width=25, height=15, 
        fg_color="#f56a6a", hover_color = "#ee4242", border_color="#ff0000",corner_radius=5,border_width=1, command=self.parar_calibrado_flujo,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_parar_calibrado_flujo.grid(row=0, column=2, ipadx = 0 ,padx=0, pady=0, sticky="nswe")



        #Calibración Concentración
        self.l_calibrado_concentracion = ctk.CTkLabel(self.f_calibracion_scroll, text="Concentración", text_color="#458B8D", font=ctk.CTkFont(size=20, weight="bold"))
        self.l_calibrado_concentracion.grid(row=3, column=0, padx=60, pady=0, sticky="ws")

        #Gráfica en tiempo real - Concentración
        fig_concentracion = Figure(figsize=(10,5),dpi=60)
        self.ax_concentracion = fig_concentracion.add_subplot(111)
        self.ax_concentracion.set_xlabel("Tiempo (s)")
        self.ax_concentracion.set_ylabel("Concentracion (µg/m\u00B3)")
        self.ax_concentracion.plot(self.tiempo_grafica_concentracion,self.buffer_concentracion)
        fig_concentracion.set_facecolor('#242424')  # Fondo de la figura transparente
        self.ax_concentracion.set_facecolor("#242424")  # Fondo del gráfico transparente
        self.ax_concentracion.tick_params(colors = "#ffffff") 
        self.ax_concentracion.xaxis.label.set_color("#ffffff")  # Color de las etiquetas del eje x
        self.ax_concentracion.yaxis.label.set_color("#ffffff")  # Color de las etiquetas del eje y
        self.ax_concentracion.spines['bottom'].set_color('#ffffff')  # Color de la línea del eje x
        self.ax_concentracion.spines['left'].set_color('#ffffff')  # Color de
        self.ax_concentracion.spines['top'].set_color('#ffffff')  # Color de la línea superior
        self.ax_concentracion.spines['right'].set_color('#ffffff')  # Color de la línea derecha
        fig_concentracion.tight_layout()
        self.canvas_concentracion = FigureCanvasTkAgg(fig_concentracion, master=self.f_calibracion_scroll)
        self.canvas_concentracion.get_tk_widget().grid(row=4, column=0, padx=10, pady=0, sticky="nsew")
        self.canvas_concentracion.draw()

        self.f_botones_calibrado_concentracion = ctk.CTkFrame(self.f_calibracion_scroll, fg_color="transparent", width=75, height=15, corner_radius=10,border_width=1)
        self.f_botones_calibrado_concentracion.grid(row=3, column=0, padx=10, pady=0, sticky="se")
        self.f_botones_calibrado_concentracion.grid_columnconfigure(0, weight=1)
        self.f_botones_calibrado_concentracion.grid_columnconfigure(1, weight=1)
        self.f_botones_calibrado_concentracion.grid_columnconfigure(2, weight=1)

        self.b_iniciar_calibrado_concentracion = ctk.CTkButton(self.f_botones_calibrado_concentracion, text="▶", text_color = "#ffffff", fg_color="#85ad75", width=25, height=15, 
        hover_color="#488f51", border_color="#006400", border_width=1, corner_radius=5,command=self.iniciar_calibrado_concentracion,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_iniciar_calibrado_concentracion.grid(row=0, column=0, ipadx = 0, padx=0, pady=0, sticky="nswe")
       

        self.b_reiniciar_calibrado_concentracion = ctk.CTkButton(self.f_botones_calibrado_concentracion, text="↻" , text_color = "#ffffff", fg_color="#C7BE19",width=25, height=15,
                                                   hover_color="#9da31e", border_color="#CFC61B", border_width=1, corner_radius=5,command=self.reiniciar_calibrado_concentracion,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_reiniciar_calibrado_concentracion.grid(row=0, column=1, ipadx = 0, padx=0, pady=0, sticky="nswe")

        self.b_parar_calibrado_concentracion = ctk.CTkButton(self.f_botones_calibrado_concentracion, text="◼", text_color = "#ffffff",width=25, height=15, 
        fg_color="#f56a6a", hover_color = "#ee4242", border_color="#ff0000",corner_radius=5,border_width=1, command=self.parar_calibrado_concentracion,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_parar_calibrado_concentracion.grid(row=0, column=2, ipadx = 0 ,padx=0, pady=0, sticky="nswe")




        #Calibración Latencia
        self.l_calibrado_latencia = ctk.CTkLabel(self.f_calibracion_scroll, text="Latencia", text_color="#458B8D", font=ctk.CTkFont(size=20, weight="bold"))
        self.l_calibrado_latencia.grid(row=3, column=1, padx=60, pady=0, sticky="ws")

        #Gráfica en tiempo real - Latencia
        fig_latencia = Figure(figsize=(10,5),dpi=60)
        self.ax_latencia = fig_latencia.add_subplot(111)
        self.ax_latencia.set_xlabel("Tiempo (s)")
        self.ax_latencia.set_ylabel("Latencia (ms)")
        self.ax_latencia.plot(self.tiempo_grafica_latencia,self.buffer_latencia)
        fig_latencia.set_facecolor('#242424')  # Fondo de la figura transparente
        self.ax_latencia.set_facecolor("#242424")  # Fondo del gráfico transparente
        self.ax_latencia.tick_params(colors = "#ffffff") 
        self.ax_latencia.xaxis.label.set_color("#ffffff")  # Color de las etiquetas del eje x
        self.ax_latencia.yaxis.label.set_color("#ffffff")  # Color de las etiquetas del eje y
        self.ax_latencia.spines['bottom'].set_color('#ffffff')  # Color de la línea del eje x
        self.ax_latencia.spines['left'].set_color('#ffffff')  # Color de
        self.ax_latencia.spines['top'].set_color('#ffffff')  # Color de la línea superior
        self.ax_latencia.spines['right'].set_color('#ffffff')  # Color de la línea derecha
        fig_latencia.tight_layout()
        self.canvas_latencia = FigureCanvasTkAgg(fig_latencia, master=self.f_calibracion_scroll)
        self.canvas_latencia.get_tk_widget().grid(row=4, column=1, padx=10, pady=0, sticky="nsew")
        self.canvas_latencia.draw()

        self.f_botones_calibrado_latencia = ctk.CTkFrame(self.f_calibracion_scroll, fg_color="transparent", width=75, height=15, corner_radius=10,border_width=1)
        self.f_botones_calibrado_latencia.grid(row=3, column=1, padx=10, pady=0, sticky="se")
        self.f_botones_calibrado_latencia.grid_columnconfigure(0, weight=1)
        self.f_botones_calibrado_latencia.grid_columnconfigure(1, weight=1)
        self.f_botones_calibrado_latencia.grid_columnconfigure(2, weight=1)

        self.b_iniciar_calibrado_latencia = ctk.CTkButton(self.f_botones_calibrado_latencia, text="▶", text_color = "#ffffff", fg_color="#85ad75", width=25, height=15, 
        hover_color="#488f51", border_color="#006400", border_width=1, corner_radius=5,command=self.iniciar_calibrado_latencia,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_iniciar_calibrado_latencia.grid(row=0, column=0, ipadx = 0, padx=0, pady=0, sticky="nswe")
       

        self.b_reiniciar_calibrado_latencia = ctk.CTkButton(self.f_botones_calibrado_latencia, text="↻" , text_color = "#ffffff", fg_color="#C7BE19",width=25, height=15,
                                                   hover_color="#9da31e", border_color="#CFC61B", border_width=1, corner_radius=5,command=self.reiniciar_calibrado_latencia,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_reiniciar_calibrado_latencia.grid(row=0, column=1, ipadx = 0, padx=0, pady=0, sticky="nswe")

        self.b_parar_calibrado_latencia = ctk.CTkButton(self.f_botones_calibrado_latencia, text="◼", text_color = "#ffffff",width=25, height=15, 
        fg_color="#f56a6a", hover_color = "#ee4242", border_color="#ff0000",corner_radius=5,border_width=1, command=self.parar_calibrado_latencia,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_parar_calibrado_latencia.grid(row=0, column=2, ipadx = 0 ,padx=0, pady=0, sticky="nswe")



        self.actualizar_graficas()


        """
        self.f_canales = ctk.CTkScrollableFrame(self,fg_color="transparent")  # Fondo transparente para el marco
        self.f_canales.grid(row=1,column=0,padx=5,pady=10,sticky="nsew",rowspan=2)  # Usamos grid para colocar el marco
        self.f_canales.grid_columnconfigure((0,1), weight=1)
        self.f_canales.grid_rowconfigure((0,1), weight=1)


        self.l_canales= ctk.CTkLabel(self.f_canales, text="Canales",text_color="#458B8D", font=ctk.CTkFont(size=22, weight="bold"))
        self.l_canales.grid(row=0, column=0, padx=10, pady=(10,5), sticky="w")

        self.cuadros_canales = []
        for i in range(0,6):
            columna = i//3
            fila = i%3
            cuadro_canal= Canal(self.f_canales,color_canal=self.colores_canales[i],num_canal=i,registro=self.consola.registro, actualizar_canal=self.actualizar_canales)
            cuadro_canal.grid(row=fila+1, column=columna, padx=10, pady=10, sticky="nsew")
            self.cuadros_canales.append(cuadro_canal)
        
        """
        self.tv_prot_cal_est = ctk.CTkTabview(master=self, fg_color="transparent",border_color="#4a4c4e", border_width=1, corner_radius=10, width=400)
        self.tv_prot_cal_est.grid(row=1,column=1,padx=5,pady=10,sticky="nsew", rowspan=3)
        self.tv_prot_cal_est.add("Protocolo")
        #self.tv_prot_cal_est.add("Calibración")
        self.tv_prot_cal_est.add("Estado")
        self.tv_prot_cal_est._segmented_button.configure(text_color="#ffffff", border_width=1, corner_radius=10, width = 200,height =30, font=ctk.CTkFont(size=20, weight="bold"))

        self.tv_prot_cal_est.tab("Protocolo").grid_columnconfigure(0, weight=1)
        #self.tv_prot_cal_est.tab("Calibración").grid_columnconfigure(0, weight=1)
        self.tv_prot_cal_est.tab("Estado").grid_columnconfigure(0, weight=1)

        self.tv_prot_cal_est.set("Protocolo")

        #self.b_protocolo = ctk.CTkButton(master=self.tv_prot_cal_est.tab("Protocolo"), text="Protocolo")
        #self.b_protocolo.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")



        
        #PROTOCOLO 
        #self.f_protocolo = ctk.CTkFrame(self, fg_color="#1e1e1e",border_color="#4a4c4e", border_width=1, corner_radius=10)   
        #self.f_protocolo.grid(row=1,column=1,padx=5,pady=10,sticky="nsew")
        #self.f_protocolo.grid_columnconfigure(0, weight=1)

        #self.l_protocolo= ctk.CTkLabel(self.tv_prot_cal_est.tab("Protocolo"), bg_color= "transparent",text="Protocolo de Olfatometría", text_color="#458B8D", 
                                       #font=ctk.CTkFont(size=22, weight="bold"))
        #self.l_protocolo.grid(row=0, column=0, padx=5, pady=(20,5), sticky="we")

        #Número de Ciclos
        self.l_num_ciclos = ctk.CTkLabel(self.tv_prot_cal_est.tab("Protocolo"), text="Número de ciclos", font=ctk.CTkFont(size=14,weight="bold"))
        self.l_num_ciclos.grid(row=0, column=0, padx=10, pady=(10,2), sticky="we", rowspan=1)
        self.e_num_ciclos =SpinboxCTk(self.tv_prot_cal_est.tab("Protocolo"), valor= 3)
        self.e_num_ciclos.grid(row=1, column=0, padx=5, pady=(4,2), sticky="n")

        #Intervalo entreciclos
        self.l_intervalo_ciclos = ctk.CTkLabel(self.tv_prot_cal_est.tab("Protocolo"), text="Intervalo entre ciclos (en sec)", font=ctk.CTkFont(size=14,weight="bold"))
        self.l_intervalo_ciclos.grid(row=2, column=0, padx=10, pady=(4,2), sticky="we", rowspan=1)
        self.e_intervalo_ciclos =SpinboxCTk(self.tv_prot_cal_est.tab("Protocolo"), valor= 60)
        self.e_intervalo_ciclos.grid(row=3, column=0, padx=5, pady=(4,2), sticky="n")

        #Tiempo de Exposición
        self.l_tiempo_exposicion = ctk.CTkLabel(self.tv_prot_cal_est.tab("Protocolo"), text="Exposición (en sec)", font=ctk.CTkFont(size=14,weight="bold"))
        self.l_tiempo_exposicion.grid(row=4, column=0, padx=10, pady=(4,2), sticky="we")
        self.e_tiempo_exposicion = SpinboxCTk(self.tv_prot_cal_est.tab("Protocolo"), valor=3)
        self.e_tiempo_exposicion.grid(row=5, column=0, padx=5, pady=(4,2), sticky="n")

        #Tiempo de Desensibilización
        self.l_tiempo_desensibilizacion = ctk.CTkLabel(self.tv_prot_cal_est.tab("Protocolo"), text="Desensibilización (en sec)", font=ctk.CTkFont(size=14,weight="bold"))
        self.l_tiempo_desensibilizacion.grid(row=6, column=0, padx=10, pady=(4,2), sticky="we")
        self.e_tiempo_desensibilizacion = SpinboxCTk(self.tv_prot_cal_est.tab("Protocolo"), valor=30)
        self.e_tiempo_desensibilizacion.grid(row=7, column=0, padx=5, pady=(4,2), sticky="n")

        #Orden de los canales
        self.l_orden_canales = ctk.CTkLabel(self.tv_prot_cal_est.tab("Protocolo"), text="Orden de los canales", font=ctk.CTkFont(size=14,weight="bold"))
        self.l_orden_canales.grid(row=8, column=0, padx=10, pady=(30,5), sticky="n")
        self.cb_secuencial = ctk.CTkCheckBox(self.tv_prot_cal_est.tab("Protocolo"), text="Secuencial", font=ctk.CTkFont(size=14))
        self.cb_secuencial.grid(row=9, column=0, padx=40, pady=(4,2), sticky="w")
        self.cb_secuencial.select()  # Por defecto, el orden de los canales es secuencial
        self.cb_aleatorio = ctk.CTkCheckBox(self.tv_prot_cal_est.tab("Protocolo"), text="Aleatorio", font=ctk.CTkFont(size=14))
        self.cb_aleatorio.grid(row=9, column=0, padx=40, pady=(4,2), sticky="e")





        #PROTOCOLOS CONOCIDOS/ESTANDARIZADOS 
        #OJO: ME QUEDA MUY LARGO SI NO PUEDO UTILIZAR LA CLASE CTkComboBox DE CUSTOM TKINTER
        #Sniffin Sticks test(identificación, umbral y descriminación)
        #self.comb_protocolos_definidos = stk.def combobox_callback(choice):
            #print('combobox dropdown clicked:', choice)
        """
        self.l_protocolos_definidos= ctk.CTkLabel(self.tv_prot_cal_est.tab("Protocolo"), text="Protocolos definidos", text_color="#ffffff", font=ctk.CTkFont(size=14, weight="bold"))
        self.l_protocolos_definidos.grid(row=11, column=0, padx=10, pady=(20,5), sticky="w")
        sv_protolocos_definidos = ctk.StringVar(value='------')
        comb_protocolos_definidos = ctk.CTkComboBox(self.tv_prot_cal_est.tab("Protocolo"), values=['------','Sniffin Sticks', 'UPSIT', 'CCCRC', 'Snap & Sniff Threshold', 'Q-Sticks', 'Protocolo papers Ángela'],
                                             width=100, height=28, text_color="#ffffff",dropdown_fg_color="#0a5f70", command= self.protocolo_definido, variable=sv_protolocos_definidos)
        comb_protocolos_definidos.grid(row=11, column=0, padx=10, pady=(20,5), sticky="e")
        """
    
        #Inicar protocolo
        self.b_iniciar_protocolo = ctk.CTkButton(self.tv_prot_cal_est.tab("Protocolo"), text="▶", text_color = "#ffffff", fg_color="#85ad75", width=15, height=20, 
        hover_color="#488f51", border_color="#006400", border_width=1, corner_radius=10,command=self.iniciar_protocolo,font=ctk.CTkFont(size=34, weight="bold"))
        self.b_iniciar_protocolo.grid(row=18, column=0, padx=50, pady=(60,5), sticky="w")
        
        #Reiniciar protocolo
        self.b_reiniciar_protocolo = ctk.CTkButton(self.tv_prot_cal_est.tab("Protocolo"), text="↻" , text_color = "#ffffff", fg_color="#C7BE19",width=15, height=20, 
                                                   hover_color="#9da31e", border_color="#F6F04F", border_width=1, corner_radius=10,command=self.reiniciar_protocolo,font=ctk.CTkFont(size=34, weight="bold"))
        self.b_reiniciar_protocolo.grid(row=18, column=0, padx=10, pady=(60,5), sticky="n")


        #Parar protocolo
        self.b_parar_protocolo = ctk.CTkButton(self.tv_prot_cal_est.tab("Protocolo"), text="◼", text_color = "#ffffff",width=15, height=20, 
        fg_color="#f56a6a", hover_color = "#ee4242", border_color="#ff0000",corner_radius=10,border_width=1, command=self.parar_protocolo,font=ctk.CTkFont(size=34, weight="bold"))
        self.b_parar_protocolo.grid(row=18, column=0, padx=50, pady=(60,5), sticky="e")
        

        #FALTAN INLCUIR PROTOCOLOS ESTANDARIZADOS DE OLFATOMETRÍA (Sniffin' Sticks, etc) Y LA POSIBILIDAD DE CREAR 
        # PROTOCOLOS PERSONALIZADOS

        """
        #CALIBRACIÓN
        # Crea un frame deslizante dentro del tabview
        self.f_calibracion_scroll = ctk.CTkScrollableFrame(self.tv_prot_cal_est.tab("Calibración"), fg_color="transparent")
        self.f_calibracion_scroll.grid(row=0, column=0, padx=0, pady=0, sticky="nsew")
        self.tv_prot_cal_est.tab("Calibración").grid_rowconfigure(0, weight=1)
        self.tv_prot_cal_est.tab("Calibración").grid_columnconfigure(0, weight=1)
        self.f_calibracion_scroll.grid_columnconfigure(0, weight=1)
        
        self.l_calibracion = ctk.CTkLabel(self.f_calibracion_scroll, text="Calibración", text_color="#458B8D", font=ctk.CTkFont(size=22, weight="bold"))
        self.l_calibracion.grid(row=0, column=0, padx=10, pady=(10,5), sticky="n")

        # Selección de canales a calibrar
        sv_canales_calibrado = ctk.StringVar(value="Ninguno")
        comb_canales_calibrados = ctk.CTkComboBox(self.f_calibracion_scroll, values=['Ninguno','Canal Amarillo', 'Canal Rojo', 'Canal Verde', 'Canal Azul', 'Canal Blanco', 'Canal Negro'],
                                             width=160, height=36, text_color="#ffffff",command=self.canal_calibrado,dropdown_fg_color="#0a5f70", variable=sv_canales_calibrado)
        comb_canales_calibrados.grid(row=1, column=0, padx=10, pady=(20,5), sticky="n")

        #Calibración Velocidad
        self.l_calibracion_velocidad = ctk.CTkLabel(self.f_calibracion_scroll, text="Velocidad", text_color="#458B8D", font=ctk.CTkFont(size=18, weight="bold"))
        self.l_calibracion_velocidad.grid(row=2, column=0, padx=10, pady=(50,5), sticky="n")

        #Gráfica en tiempo real - Velocidad
        fig_velocidad = Figure(figsize=(10,5),dpi=55)
        self.ax_velocidad = fig_velocidad.add_subplot(111)
        self.ax_velocidad.set_xlabel("Tiempo (s)")
        self.ax_velocidad.set_ylabel("Velocidad (m/s)")
        self.ax_velocidad.plot(self.tiempo_graficas,self.buffer_velocidad)
        fig_velocidad.tight_layout()
        self.canvas_velocidad = FigureCanvasTkAgg(fig_velocidad, master=self.f_calibracion_scroll)
        self.canvas_velocidad.get_tk_widget().grid(row=3, column=0, padx=10, pady=20, sticky="n")
        self.canvas_velocidad.draw()

        self.b_iniciar_calibracion_velocidad = ctk.CTkButton(self.f_calibracion_scroll, text="▶", text_color = "#ffffff", fg_color="#85ad75", width=15, height=20, 
        hover_color="#488f51", border_color="#006400", border_width=1, corner_radius=10,command=self.iniciar_calibrado_velocidad,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_iniciar_calibracion_velocidad.grid(row=4, column=0, padx=90, pady=(20,5), sticky="w")
       

        self.b_reiniciar_calibrado_velocidad = ctk.CTkButton(self.f_calibracion_scroll, text="↻" , text_color = "#ffffff", fg_color="#C7BE19",width=15, height=20,
                                                   hover_color="#9da31e", border_color="#F6F04F", border_width=1, corner_radius=10,command=self.reiniciar_calibrado_velocidad,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_reiniciar_calibrado_velocidad.grid(row=4, column=0, padx=10, pady=(20,5), sticky="n")

        self.b_parar_calibrado_velocidad = ctk.CTkButton(self.f_calibracion_scroll, text="◼", text_color = "#ffffff",width=15, height=20, 
        fg_color="#f56a6a", hover_color = "#ee4242", border_color="#ff0000",corner_radius=10,border_width=1, command=self.parar_calibrado_velocidad,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_parar_calibrado_velocidad.grid(row=4, column=0, padx=90, pady=(20,5), sticky="e")


        #Calibración Flujo
        self.l_calibracion_flujo = ctk.CTkLabel(self.f_calibracion_scroll, text="Flujo", text_color="#458B8D", font=ctk.CTkFont(size=18, weight="bold"))
        self.l_calibracion_flujo.grid(row=5, column=0, padx=10, pady=(50,5), sticky="n")

        #Gráfica en tiempo real - Flujo
        fig_flujo = Figure(figsize=(10,5),dpi=60)
        self.ax_flujo = fig_flujo.add_subplot(111)
        self.ax_flujo.set_xlabel("Tiempo (s)")
        self.ax_flujo.set_ylabel("Flujo (ml/min)")
        self.ax_flujo.plot(self.tiempo_graficas,self.buffer_flujo)
        fig_flujo.tight_layout()
        self.canvas_flujo = FigureCanvasTkAgg(fig_flujo, master=self.f_calibracion_scroll)
        self.canvas_flujo.get_tk_widget().grid(row=6, column=0, padx=10, pady=(20,5), sticky="n")
        self.canvas_flujo.draw()

        self.b_iniciar_calibracion_flujo = ctk.CTkButton(self.f_calibracion_scroll, text="▶", text_color = "#ffffff", fg_color="#85ad75", width=15, height=20, 
        hover_color="#488f51", border_color="#006400", border_width=1, corner_radius=10,command=self.iniciar_calibrado_flujo,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_iniciar_calibracion_flujo.grid(row=7, column=0, padx=90, pady=(20,5), sticky="w")
       

        self.b_reiniciar_calibrado_flujo = ctk.CTkButton(self.f_calibracion_scroll, text="↻" , text_color = "#ffffff", fg_color="#C7BE19",width=15, height=20,
                                                   hover_color="#9da31e", border_color="#F6F04F", border_width=1, corner_radius=10,command=self.reiniciar_calibrado_flujo,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_reiniciar_calibrado_flujo.grid(row=7, column=0, padx=10, pady=(20,5), sticky="n")

        self.b_parar_calibrado_flujo = ctk.CTkButton(self.f_calibracion_scroll, text="◼", text_color = "#ffffff",width=15, height=20, 
        fg_color="#f56a6a", hover_color = "#ee4242", border_color="#ff0000",corner_radius=10,border_width=1, command=self.parar_calibrado_flujo,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_parar_calibrado_flujo.grid(row=7, column=0, padx=90, pady=(20,5), sticky="e")



        #Calibración Concentración
        self.l_calibracion_concentracion = ctk.CTkLabel(self.f_calibracion_scroll, text="Concentración", text_color="#458B8D", font=ctk.CTkFont(size=18, weight="bold"))
        self.l_calibracion_concentracion.grid(row=8, column=0, padx=10, pady=(50,5), sticky="n")

        #Gráfica en tiempo real - Concentración
        fig_concentracion = Figure(figsize=(10,5),dpi=60)
        self.ax_concentracion = fig_concentracion.add_subplot(111)
        self.ax_concentracion.set_xlabel("Tiempo (s)")
        self.ax_concentracion.set_ylabel("Concentracion (µg/m\u00B3)")
        self.ax_concentracion.plot(self.tiempo_graficas,self.buffer_concentracion)
        fig_concentracion.tight_layout()
        self.canvas_concentracion = FigureCanvasTkAgg(fig_concentracion, master=self.f_calibracion_scroll)
        self.canvas_concentracion.get_tk_widget().grid(row=9, column=0, padx=10, pady=(20,5), sticky="n")
        self.canvas_concentracion.draw()

        self.b_iniciar_calibracion_concentracion = ctk.CTkButton(self.f_calibracion_scroll, text="▶", text_color = "#ffffff", fg_color="#85ad75", width=15, height=20, 
        hover_color="#488f51", border_color="#006400", border_width=1, corner_radius=10,command=self.iniciar_calibrado_concentracion,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_iniciar_calibracion_concentracion.grid(row=10, column=0, padx=90, pady=(20,5), sticky="w")
       

        self.b_reiniciar_calibrado_concentracion = ctk.CTkButton(self.f_calibracion_scroll, text="↻" , text_color = "#ffffff", fg_color="#C7BE19",width=15, height=20,
                                                   hover_color="#9da31e", border_color="#F6F04F", border_width=1, corner_radius=10,command=self.reiniciar_calibrado_concentracion, font=ctk.CTkFont(size=24, weight="bold"))
        self.b_reiniciar_calibrado_concentracion.grid(row=10, column=0, padx=10, pady=(20,5), sticky="n")

        self.b_parar_calibrado_concentracion = ctk.CTkButton(self.f_calibracion_scroll, text="◼", text_color = "#ffffff",width=15, height=20, 
        fg_color="#f56a6a", hover_color = "#ee4242", border_color="#ff0000",corner_radius=10,border_width=1, command=self.parar_calibrado_concentracion,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_parar_calibrado_concentracion.grid(row=10, column=0, padx=90, pady=(20,5), sticky="e")




        #Calibración Latencia
        self.l_calibracion_latencia = ctk.CTkLabel(self.f_calibracion_scroll, text="Latencia", text_color="#458B8D", font=ctk.CTkFont(size=18, weight="bold"))
        self.l_calibracion_latencia.grid(row=11, column=0, padx=10, pady=(50,5), sticky="n")

        #Gráfica en tiempo real - Latencia
        fig_latencia = Figure(figsize=(10,5),dpi=60)
        self.ax_latencia = fig_latencia.add_subplot(111)
        self.ax_concentracion.set_xlabel("Tiempo (s)")
        self.ax_concentracion.set_ylabel("Latencia (ms)")
        fig_latencia.tight_layout()
        self.canvas_latencia = FigureCanvasTkAgg(fig_latencia, master=self.f_calibracion_scroll)
        self.canvas_latencia.get_tk_widget().grid(row=12, column=0, padx=10, pady=(20,5), sticky="n")
        self.canvas_latencia.draw()

        self.b_iniciar_calibracion_latencia = ctk.CTkButton(self.f_calibracion_scroll, text="▶", text_color = "#ffffff", fg_color="#85ad75", width=15, height=20, 
        hover_color="#488f51", border_color="#006400", border_width=1, corner_radius=10,command=self.iniciar_calibrado_latencia,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_iniciar_calibracion_latencia.grid(row=13, column=0, padx=90, pady=(20,5), sticky="w")
       

        self.b_reiniciar_calibrado_latencia = ctk.CTkButton(self.f_calibracion_scroll, text="↻" , text_color = "#ffffff", fg_color="#C7BE19",width=15, height=20,
                                                   hover_color="#9da31e", border_color="#F6F04F", border_width=1, corner_radius=10,command=self.reiniciar_calibrado_latencia, font=ctk.CTkFont(size=24, weight="bold"))
        self.b_reiniciar_calibrado_latencia.grid(row=13, column=0, padx=10, pady=(20,5), sticky="n")

        self.b_parar_calibrado_latencia = ctk.CTkButton(self.f_calibracion_scroll, text="◼", text_color = "#ffffff",width=15, height=20, 
        fg_color="#f56a6a", hover_color = "#ee4242", border_color="#ff0000",corner_radius=10,border_width=1, command=self.parar_calibrado_latencia,font=ctk.CTkFont(size=24, weight="bold"))
        self.b_parar_calibrado_latencia.grid(row=13, column=0, padx=90, pady=(20,5), sticky="e")



        self.actualizar_graficas()



        self.b_iniciar_calibracion = ctk.CTkButton(self.f_calibracion_scroll, text="Iniciar Calibrado General", text_color="#ffffff", fg_color= "#85ad75", width= 15, height=20,
                                                   hover_color="#488f51", border_color="#006400", border_width=1, corner_radius=10,command=self.iniciar_calibrado,font=ctk.CTkFont(size=20, weight="bold"))
        self.b_iniciar_calibracion.grid(row=14, column=0, padx=10, pady=(70,5), sticky="n")

        
        self.b_reiniciar_calibrado = ctk.CTkButton(self.f_calibracion_scroll, text="Reiniciar Calibrado General", text_color="#ffffff", fg_color= "#C7BE19" , width=15, height=20,
                                                   hover_color="#9da31e", border_color="#F6F04F", border_width= 1, corner_radius=10,command=self.reiniciar_calibrado,font=ctk.CTkFont(size=20, weight="bold"))
        self.b_reiniciar_calibrado.grid(row=15, column=0, padx=10, pady=(20,5), sticky="n")


        self.b_parar_calibrado = ctk.CTkButton(self.f_calibracion_scroll, text="Parar Calibrado General", text_color="#ffffff", fg_color= "#f56a6a", width=15, height=20,
                                                   hover_color="#ee4242", border_color="#ff0000", border_width= 1, corner_radius=10,command=self.parar_calibrado,font=ctk.CTkFont(size=20, weight="bold"))
        self.b_parar_calibrado.grid(row=16, column=0, padx=10, pady=(20,5), sticky="n")

        """



        #ESTADO
        #self.f_estado = ctk.CTkFrame(self, fg_color="#1e1e1e",border_color="#4a4c4e", border_width=1, corner_radius=10)
        #self.f_estado.grid(row=3,column=1,padx=5,pady=(0,10),sticky="nsew")
        #self.columnconfigure(0, weight=1)


        #self.l_titulo_estado= ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), text="Estado", text_color="#458B8D", font=ctk.CTkFont(size=22, weight="bold"))
        #self.l_titulo_estado.grid(row=0, column=0, padx=10, pady=(10,5), sticky="n")

        self.l_fecha = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), text=f"Fecha y hora: {time.strftime('%d-%m-%Y %H:%M:%S', time.localtime())}", font=ctk.CTkFont(size=14, weight="bold"))
        self.l_fecha.grid(row=0, column=0, padx=10, pady=(20,5), sticky="n")

        
        self.tiempo_inicio_sesion = time.time()  # Variable para almacenar el tiempo de inicio de la sesión
        self.duracion_sesion= ctk.StringVar(value="Duración de sesión: 00:00:00")  # Variable para almacenar la duración de la sesión
        self.l_duracion_sesion = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), textvariable=self.duracion_sesion, font=ctk.CTkFont(size=14, weight="bold"))
        self.l_duracion_sesion.grid(row=1, column=0, padx=10, pady=(20,5), sticky="n")

        self.l_id_sesion = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), text="ID de sesión: ", font=ctk.CTkFont(size=14, weight="bold"))
        self.l_id_sesion.grid(row=2, column=0, padx=71, pady=(20,5), sticky="w")
        self.e_id_sesion= ctk.CTkEntry(self.tv_prot_cal_est.tab("Estado"), placeholder_text="introduzca identificador", font=ctk.CTkFont(size=14))
        self.e_id_sesion.grid(row=2, column=0, padx=71, pady=(20,5), sticky="e")

        self.l_id_paciente = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), text="ID de paciente: ", font=ctk.CTkFont(size=14, weight="bold"))
        self.l_id_paciente.grid(row=3, column=0, padx=65, pady=(20,5), sticky="w")
        self.e_id_paciente= ctk.CTkEntry(self.tv_prot_cal_est.tab("Estado"), placeholder_text="introduzca identificador", font=ctk.CTkFont(size=14))
        self.e_id_paciente.grid(row=3, column=0, padx=65, pady=(20,5), sticky="e")


        #self.actualizar_canales()

        self.l_canal_anterior= ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), text="Canal anterior: ", font=ctk.CTkFont(size=14, weight="bold"))
        self.l_canal_anterior.grid(row=4, column=0, padx=110, pady=(20,5), sticky="w")
        #self.canal_anterior = ctk.StringVar(value="Ninguno")
        self.l_canal_anterior_valor = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), textvariable=self.sv_canal_anterior, font=ctk.CTkFont(size=14))
        self.l_canal_anterior_valor.grid(row=4, column=0, padx=110, pady=(20,5), sticky="e")

        self.l_canal_activo= ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), text="Canal activo: ", font=ctk.CTkFont(size=14, weight="bold"))
        self.l_canal_activo.grid(row=5, column=0, padx=115, pady=(20,5), sticky="w")

        self.l_canal_activo_valor = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), textvariable=self.sv_canal_activo, font=ctk.CTkFont(size=14))
        self.l_canal_activo_valor.grid(row=5, column=0, padx=115, pady=(20,5), sticky="e")

        self.l_canal_siguiente= ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), text="Canal siguiente: ", font=ctk.CTkFont(size=14, weight="bold"))
        self.l_canal_siguiente.grid(row=6, column=0, padx=105, pady=(20,5), sticky="w")

        self.l_canal_siguiente_valor = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), textvariable=self.sv_canal_siguiente, font=ctk.CTkFont(size=14))
        self.l_canal_siguiente_valor.grid(row=6, column=0, padx=105, pady=(20,5), sticky="e")

        self.l_latencia_canal= ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), text="Latencia: ", font=ctk.CTkFont(size=14, weight="bold"))
        self.l_latencia_canal.grid(row=7, column=0, padx=140, pady=(20,5), sticky="w")
        self.sv_latencia_canal= ctk.StringVar(value="0 ms")
        self.l_valor_latencia_canal = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), textvariable=self.sv_latencia_canal, font=ctk.CTkFont(size=14))
        self.l_valor_latencia_canal.grid(row=7, column=0, padx=140, pady=(20,5), sticky="e")

        self.l_flujo_aire_canal = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), text="Flujo de aire estimado: ", font=ctk.CTkFont(size=14, weight="bold"))
        self.l_flujo_aire_canal.grid(row=8, column=0, padx=80, pady=(20,5), sticky="w")
        self.sv_flujo_aire_canal= ctk.StringVar(value="0 ml/min")
        self.l_valor_flujo_aire_canal = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), textvariable=self.sv_flujo_aire_canal, font=ctk.CTkFont(size=14))
        self.l_valor_flujo_aire_canal.grid(row=8, column=0, padx=80, pady=(20,5), sticky="e")

        self.l_concentracion_canal = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), text="Concentración estimada: ", font=ctk.CTkFont(size=14, weight="bold"))
        self.l_concentracion_canal.grid(row=9, column=0, padx=75, pady=(20,5), sticky="w")
        self.sv_concentracion_canal= ctk.StringVar(value="0 µg/m\u00B3")
        self.l_valor_concentracion_canal = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), textvariable=self.sv_concentracion_canal, font=ctk.CTkFont(size=14))
        self.l_valor_concentracion_canal.grid(row=9, column=0, padx=75, pady=(20,5), sticky="e")   

        """
        self.l_velocidad_motor_canal = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), text="Velocidad motor: ", font=ctk.CTkFont(size=14, weight="bold"))
        self.l_velocidad_motor_canal.grid(row=11, column=0, padx=70, pady=(20,5), sticky="w")
        self.sv_velocidad_motor_canal= ctk.StringVar(value="0 rpm")
        self.l_valor_velocidad_motor_canal = ctk.CTkLabel(self.tv_prot_cal_est.tab("Estado"), textvariable=self.sv_velocidad_motor_canal, font=ctk.CTkFont(size=14))
        self.l_valor_velocidad_motor_canal.grid(row=11, column=0, padx=70, pady=(20,5), sticky="e")
        """

        
        self.b_generar_informe = ctk.CTkButton(self.tv_prot_cal_est.tab("Estado"), text="Generar informe de sesión", fg_color="#5172a4",text_color="#ffffff",
                                       corner_radius=10,border_width=1, border_color="#6ba2f4",command=self.generar_informe,
                                       font=ctk.CTkFont(size=18, weight="bold"))
        self.b_generar_informe.grid(row=10, column=0, padx=31, pady=(30,5), sticky="n")

        #self.l_dispositivo_conectado?


        #Para un botón
        """
        self.button = ctk.CTkButton(self.Frame, text="Pulsar", command=self.button_click,
                                     font= ctk.CTkFont(size=16, weight="bold"),
                                     bg_color="blue", fg_color="white", hover_color="lightblue")
        self.button.grid(row=1, column=0, pady=20, sticky="nesw")
        """
    
# ─────────────────────────────────────────────────────────────
#  FUNCIONALIDADES
# ─────────────────────────────────────────────────────────────
    #def buscqueda_esp32(self):

    def button_click(self):
        self.label.configure(text="Botón pulsado")

    def generar_informe(self):
        
        formatos = [('PDF','*.pdf'),('Excel', '*.xlsx'),('CSV','*.csv')]
        ruta = ctk.filedialog.asksaveasfilename(title='Guardar informe de sesión',
                                                filetypes=formatos, defaultextension=".pdf",
                                                  initialfile=f'informe_{self.e_id_sesion.get()}_{datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}')
        if not ruta:
            self.consola.registro('Informe no generado', nivel = "AVISO")
            return
        
        self.consola.registro(f'Generando informe en {ruta}....')

        if ruta.endswith('.pdf'):
            self.generar_pdf(ruta)
        elif ruta.endswith('.xlsx'):
            self.generar_excel(ruta)
        elif ruta.endswith('.csv'):
            self.generar_csv(ruta)

        self.consola.registro(f'Informe generado en: {ruta}')

    def generar_csv(self,ruta):
        seccion_sesion=[["OlfaMetric - Informe de sesión"],
                         ["ID Sesión", self.e_id_sesion.get()],
                         ["ID Paciente", self.e_id_paciente.get()],
                         ["Fecha",datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                         ["Duración sesión", self.duracion_sesion.get().replace("Duración de sesión: ","")],
                         []]#fila vacía como separador
        
        seccion_sesion.append([])
        
        seccion_protocolo=[["PARÁMETROS DEL PROTOCOLO"],
                            ["Número de ciclos", self.e_num_ciclos.get()],
                            ["Tiempo de exposición", self.e_tiempo_exposicion.get()],
                            ["Tiempo de desensibilización", self.e_tiempo_desensibilizacion.get()],
                            ["Tiempo de intervalo entre ciclos", self.e_intervalo_ciclos.get()],
                            []
                            ]
        
        seccion_protocolo.append([])

        seccion_calibrado = [["RESULTADOS CALIBRACIÓN"],["Canal", "Olor", "Concentración bruta (µg/m\u00B3)", 
                                   "Concentración neta (µg/m\u00B3)", 
                                   "Flujo (ml/min)","Velocidad (rpm)","Latencia (ms)"]]

        #si la variable metricas_calibracion existe ejecuta el siguiente código
        if self.metricas_calibracion:
            for num_canal, metrica in self.metricas_calibracion.items():
                seccion_calibrado.append([self.colores_canales[num_canal],
                                          metrica.get("olor",""),
                                          #redondea al segundo decimal la concentración bruta medida en la calibración del canal
                                          #si la concentración bruta no existe, devuelve 0 como valor por defecto
                                          round(self.media(metrica.get("concentracion", 0)), 2),
                                          round(self.media(metrica.get("concentracion bruta", 0)), 2),
                                          round(self.media(metrica.get("concentracion neta", 0)), 2),
                                          round(self.media(metrica.get("flujo", 0)), 2),
                                          round(self.media(metrica.get("velocidad",0)), 1),
                                          round(self.media(metrica.get("latencia",0)), 1)])
                
        seccion_calibrado.append([])

        #historial completo de los datos
        seccion_historial = [["HISTORIAL DE DATOS"],["Hora","Canal", "Olor",
                                   "Estado","Flujo (ml/min)","Concentración (µg/m\u00B3)",
                                   "Concentración bruta (µg/m\u00B3),",
                                   "Concentración neta (µg/m\u00B3)", "Velocidad (rpm)","Latencia (ms)"]]
        
        for metrica in self.historial_sesion :
            #las claves definidas deben coincidir con las del simulador
            seccion_historial.append([
                datetime.datetime.fromtimestamp(metrica["timestamp"]).strftime("%H:%M:%S"),
                self.colores_canales[metrica["canal"]],
                metrica.get("olor",""),
                metrica.get("estado",""),
                metrica.get("flujo",0),
                metrica.get("concentracion"),
                metrica.get("concentracion bruta",0),
                metrica.get("concentracion neta",0),
                metrica.get("velocidad_motor",0),
                metrica.get("latencia",0)
            ])
            
        
        
        try:
            #Se añade el BOM de UTF-8 al principio del archivo. Sin esto Excel en Windows 
            # no muestra correctamente los caracteres especiales como µ o ó.
            with open(ruta, "w", newline="", encoding="utf-8-sig") as fila:
                writer = csv.writer(fila, delimiter=";")
                
                writer.writerows(seccion_sesion)
                writer.writerows(seccion_protocolo)
                writer.writerows(seccion_calibrado)
                writer.writerows(seccion_historial)

            self.consola.registro("CSV guardado correctamente")
        
        except Exception as e:
            self.consola.registro(f"Error al guardar CSV: {e}", nivel= "ERROR")

    def generar_excel(self,ruta):
        workbook = openpyxl.Workbook()

        #Estilos
        estilo_titulo = Font(bold=True, size=14, color='FF01BDCE')
        estilo_cabecera = Font(bold=True, color='FFFFFFFF')
        relleno_cabecera = PatternFill('solid', fgColor='FF1E1E1E')
        #relleno_fila_par = PatternFill('solid', fgColor='FF2B2B2B')
        centrado = Alignment(horizontal='center', vertical='center')

        def aplicar_estilo_cabecera(celda,texto):
            celda.value = texto
            celda.font = estilo_cabecera
            celda.fill = relleno_cabecera
            celda.alignment = centrado

        ##HOJA 1: RESUMEN DE SESIÓN-----------------------------------------
        hoja_resumen = workbook.active
        hoja_resumen.title = "Resumen de sesión"

        #Datos de sesión
        hoja_resumen['A1'].value = "OlfaMetric - Informe de sesión"
        hoja_resumen['A1'].font = Font(bold=True, size=20, color='FF01BDCE')
        
        datos_sesion = [
            ("ID sesión", self.e_id_sesion.get()),
            ("ID paciente", self.e_id_paciente.get()),
            ("Fecha", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Duración de sesión", self.duracion_sesion.get().replace("Duración de sesión: ",""))
        ]
        for contador, (clave,valor) in enumerate(datos_sesion, start=2):
            hoja_resumen[f'A{contador}'] = clave
            hoja_resumen[f'A{contador}'].font = Font(bold = True)
            hoja_resumen[f'B{contador}'] = valor
        
        #Parámetros del protocolo
        fila = 7
        hoja_resumen[f'A{fila}'] = "PARÁMETROS DEL PROTOCOLO"
        hoja_resumen[f'A{fila}'].font = estilo_titulo
        fila += 1

        datos_protocolo = [
            ("Número de ciclos", self.e_num_ciclos.get()),
            ("Tiempo de exposición", f'{self.e_tiempo_exposicion.get()} s'),
            ("Tiempo de desensibilización", f'{self.e_tiempo_desensibilizacion.get()} s'),
            ("Tiempo de intervalo entre ciclos", f'{self.e_intervalo_ciclos.get()} s')
        ]

        for clave, valor in datos_protocolo:
            hoja_resumen[f'A{fila}'] = clave
            hoja_resumen[f'A{fila}'].font = Font(bold=True)
            hoja_resumen[f'B{fila}'] = valor
            fila += 1
        
        #Resultados de calibración
        fila += 1
        hoja_resumen[f'A{fila}'] = "RESULTADOS CALIBRACIÓN"
        hoja_resumen[f'A{fila}'].font = estilo_titulo
        fila +=1

        #se definen las cabeceras comunes a todos los datos
        cabeceras_calibracion = ["Canal", "Olor", "Concentración bruta (µg/m\u00B3)", 
                                "Concentración neta (µg/m\u00B3)", "Flujo medio (ml/min)",
                                "Velocidad media(rpm)","Latencia media(ms)"]
        
        #se le aplica el estilo creado previamente correspondiente a las cabeceras
        for columna, cabecera in enumerate(cabeceras_calibracion, start=1):
            aplicar_estilo_cabecera(hoja_resumen.cell(fila, columna),cabecera)

        #salto de fila
        fila +=1
        

        #en caso de existir los datos relativos a la calibración se introducen en sus correspondientes
        #celdas mediante un bucle anidado, por el que con cada fila par se aplica un relleno característico
        if self.metricas_calibracion:
            for num_canal, metrica in self.metricas_calibracion.items():
                fila_datos = [
                    self.colores_canales[num_canal],
                    metrica.get("olor",""),
                    round(self.media(metrica.get("concentracion", 0)), 2),
                    round(self.media(metrica.get("concentracion bruta", 0)),2),
                    round(self.media(metrica.get("concentracion neta", 0)),2),
                    round(self.media(metrica.get("flujo", 0)),2),
                    round(self.media(metrica.get("velocidad",0)),1),
                    round(self.media(metrica.get("latencia",0)),1)
                    ]
                for columna, valor in enumerate(fila_datos, start=1):
                    hoja_resumen.cell(fila, columna).value = valor
                    #if fila % 2 == 0:
                        #hoja_resumen.cell(fila,columna).fill = relleno_fila_par
                
                fila +=1
                    
        #Se ajusta el ancho de las columnas (por comodidad)
        #auto-sizes each worksheet column based on the widest cell content in that column.
        for columna in hoja_resumen.columns:
            max_ancho = max(len(str(celda.value or "")) for celda in columna)
            hoja_resumen.column_dimensions[columna[0].column_letter].width = max_ancho + 4

        ##HOJA 2 --------------------------------
        hoja_historial = workbook.create_sheet("Historial de datos")
        hoja_historial.title = "Historial de datos"

        cabeceras_historial_datos = ["Hora","Canal", "Olor", "Estado","Flujo (ml/min)",
                                     "Velocidad (rpm)","Concentración (µg/m\u00B3)","Concentración bruta (µg/m\u00B3)",
                                     "Concentración neta (µg/m\u00B3)", "Latencia (ms)"]
        
        for columna, cabecera in enumerate(cabeceras_historial_datos,start=1):
            aplicar_estilo_cabecera(hoja_historial.cell(1,columna),cabecera)

        for fila, metrica in enumerate(self.historial_sesion, start=2):
            fila_datos = [
                datetime.datetime.fromtimestamp(metrica["timestamp"]).strftime("%H:%M:%S"),  
                self.colores_canales[metrica["canal"]],
                metrica.get("olor",""),
                metrica.get("estado",""),
                round(self.media(metrica.get("flujo", 0)), 2),
                round(self.media(metrica.get("velocidad_motor", 0)), 1),
                round(self.media(metrica.get("concentracion", 0)), 2),
                round(self.media(metrica.get("concentracion bruta", 0)), 2),
                round(self.media(metrica.get("concentracion neta", 0)), 2),
                round(self.media(metrica.get("latencia", 0)), 1),
            ]

            for columna, valor in enumerate(fila_datos, start=1):
                hoja_historial.cell(fila, columna).value = valor
                #if fila % 2 == 0:
                    #hoja_historial.cell(fila,columna).fill = relleno_fila_par

        for columna in hoja_historial.columns:
            max_ancho = max(len(str(celda.value or "")) for celda in columna)
            hoja_historial.column_dimensions[columna[0].column_letter].width = max_ancho + 4

        ##HOJA 3 -----------------------------------
        hoja_graficas = workbook.create_sheet("Gráficas de Calibrado")
        hoja_graficas.title = "Gráficas de Calibrado"
        hoja_graficas['A1'].font = estilo_titulo

        if self.metricas_calibracion:
            fila = 3
            for num_canal, metricas in self.metricas_calibracion.items():
                color_canal = self.colores_canales[num_canal]
                #muestras = metrica.get("muestras", {})

                for parametro, valores in metricas.items():
                    #se generar grafícas mediante el paquetede matplotlib, y se guardan como imágenes temporales
                    figura, eje = plt.subplots(figsize=(8,3), dpi=80)
                    eje.plot(valores, color="#01bdce")
                    eje.set_title(f'Canal {color_canal}/{self.cuadros_canales[num_canal].e_olor_canal.get()} - {parametro}', fontdict={'fontsize': 10, 'fontweight': 'bold'})
                    eje.set_xlabel('Tiempo (s)')
                    eje.set_ylabel(parametro)
                    eje.grid(True, linestyle='--', alpha=0.5)
                    eje.set_facecolor('#2b2b2b')
                    eje.patch.set_facecolor('#1e1e1e')
                    eje.tick_params(colors="white")
                    eje.title.set_color("white")
                    eje.xaxis.label.set_color('white')
                    eje.yaxis.label.set_color('white')

                    for spine in eje.spines.values():
                        spine.set_color('white')

                    #Se guarda temporalmente la figura/gráfica generada, para luego insertarla en el Excel
                    with tempfile.NamedTemporaryFile(suffix=".png", delete= False) as temporal:
                        ruta_figura = temporal.name

                    figura.savefig(ruta_figura, bbox_inches='tight', facecolor=figura.get_facecolor())
                    plt.close(figura)

                    #Ahora se inserta la imagen en la hoja Excel de la fila correspondiente.
                    imagen = openpyxl.drawing.image.Image(ruta_figura)
                    imagen.anchor = f'A{fila}'
                    hoja_graficas.add_image(imagen)
                    fila += 18

                    os.unlink(ruta_figura)  # Elimina la imagen temporal después de insertarla en el Excel
                    self.consola.registro("LLEGA HASTA AQUÍiiiiiii")

        try:
            #guardamos el workbook con el conjunto de las hojas y datos generados en la ruta seleccionada 
            # por el usuario
            workbook.save(ruta)
            self.consola.registro("Excel guardado correctamente")

        except Exception as e:
            self.consola.registro(f"Error al guardar Excel: {e}", nivel="ERROR")


        



    def generar_pdf(self, ruta):
        import statistics
        from reportlab.lib.enums import TA_CENTER


        # Estilos
        estilos = getSampleStyleSheet()
        estilo_titulo = ParagraphStyle('Titulo',
                                    fontSize=25,
                                    textColor=colors.HexColor('#01BDCE'),
                                    alignment=TA_CENTER,          # ✅ constante, no string
                                    spaceAfter=20,
                                    fontName='Helvetica-Bold')

        estilo_seccion = ParagraphStyle('Seccion',
                                    fontSize=16,
                                    textColor=colors.HexColor('#01BDCE'),
                                    spaceBefore=15,
                                    spaceAfter=10,
                                    fontName='Helvetica-Bold')

        estilo_tabla_cabecera = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E1E1E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#EEEEEE'), colors.white]),  # ✅ ROWBACKGROUNDS con S
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('PADDING', (0, 0), (-1, -1), 8),
        ])

        historia = []

        # TÍTULO
        historia.append(Paragraph("OlfaMetric - Informe de sesion", estilo_titulo))  # ✅ sin ó
        historia.append(Spacer(1, 4*mm))

        # DATOS DE SESIÓN
        historia.append(Paragraph("Datos de sesion", estilo_seccion))
        datos_sesion = [
            ["ID sesion", self.e_id_sesion.get()],
            ["ID paciente", self.e_id_paciente.get()],
            ["Fecha", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Duracion de sesion", self.duracion_sesion.get().replace("Duración de sesión: ", "")]
        ]
        tabla_sesion = Table(datos_sesion, colWidths=(100*mm, 80*mm))
        tabla_sesion.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('PADDING', (0, 0), (-1, -1), 8),
            ]))
        
        historia.append(tabla_sesion)
        historia.append(Spacer(1, 10*mm))

        # PARÁMETROS DEL PROTOCOLO
        historia.append(Paragraph("Parametros del protocolo", estilo_seccion))
        datos_protocolo = [
            ["Numero de ciclos", self.e_num_ciclos.get()],
            ["Tiempo de exposicion", f'{self.e_tiempo_exposicion.get()} s'],
            ["Tiempo de desensibilizacion", f'{self.e_tiempo_desensibilizacion.get()} s'],
            ["Tiempo de intervalo entre ciclos", f'{self.e_intervalo_ciclos.get()} s']
        ]
        tabla_protocolo = Table(datos_protocolo, colWidths=(100*mm, 80*mm))
        tabla_protocolo.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('PADDING', (0, 0), (-1, -1), 8),
            ]))
        historia.append(tabla_protocolo)
        historia.append(Spacer(1, 10*mm))

        # RESULTADOS DE CALIBRACIÓN
        historia.append(Paragraph("Resultados de calibracion", estilo_seccion))
        cabeceras_calibracion = [[                                    # ✅ lista de listas
        "Canal", "Olor", "Conc.(ug/m³)", "C.bruta(ug/m³)",
        "C.neta(ug/m³)", "Flujo(ml/min)", "Vel (rpm)", "Latencia(ms)"
        ]]
        datos_calibracion = []
        if self.metricas_calibracion:
            for num_canal, metrica in self.metricas_calibracion.items():
                datos_calibracion.append([
                    self.colores_canales[num_canal],
                    metrica.get("olor", ""),
                    round(self.media(metrica.get("concentracion", [])), 2),
                    round(self.media(metrica.get("concentracion bruta", [])), 2),
                    round(self.media(metrica.get("concentracion neta", [])), 2),
                    round(self.media(metrica.get("flujo", [])), 2),
                    round(self.media(metrica.get("velocidad", [])), 1),
                    round(self.media(metrica.get("latencia", [])), 1),
                    ])
        tabla_calibracion = Table(cabeceras_calibracion + datos_calibracion,
                               colWidths=[22*mm, 24*mm, 26*mm, 28*mm, 28*mm, 24*mm, 22*mm, 24*mm])
        tabla_calibracion.setStyle(estilo_tabla_cabecera)
        historia.append(tabla_calibracion)
        historia.append(PageBreak())

        #HISTORIAL DE DATOS
        historia.append(Paragraph("Historial de datos", estilo_seccion))
        cabeceras_historial = [[                                      # ✅ lista de listas
        "Hora", "Canal", "Olor", "Flujo(ml/min)",
        "Vel.(rpm)", "Conc.(ug/m³)", "C.bruta(ug/m³)",
        "C.neta(ug/m³)", "Latencia(ms)"
        ]]
        datos_historial = []
        for metrica in self.historial_sesion:
            datos_historial.append([
                datetime.datetime.fromtimestamp(metrica["timestamp"]).strftime("%H:%M:%S"),
                self.colores_canales[metrica["canal"]],
                metrica.get("olor", ""),
                #metrica.get("estado", ""),
                round(metrica.get("flujo", 0), 2),
                round(metrica.get("velocidad_motor", 0), 1),
                round(metrica.get("concentracion", 0), 2),
                round(metrica.get("concentracion bruta", 0), 2),
                round(metrica.get("concentracion neta", 0), 2),
                round(metrica.get("latencia", 0), 1),
                ])
        tabla_historial = Table(cabeceras_historial + datos_historial,
                             colWidths=[16*mm, 20*mm, 22*mm, 24*mm, 20*mm, 24*mm, 26*mm, 26*mm, 24*mm])
        #                         total: 16+16+26+16+20+18+22+22+22+22 = 200mm → ajustar a 180mm
        tabla_historial.setStyle(estilo_tabla_cabecera)
        historia.append(tabla_historial)

        # GENERAR PDF
        try:
            documento = SimpleDocTemplate(ruta, pagesize=A4,
                                       rightMargin=15*mm, leftMargin=15*mm,
                                       topMargin=15*mm, bottomMargin=15*mm)
            documento.build(historia)
            self.consola.registro("PDF guardado correctamente")
        except Exception as e:
            import traceback
            self.consola.registro(f"Error al guardar PDF: {e}", nivel="ERROR")
            self.consola.registro(traceback.format_exc(), nivel="ERROR")






    def b_buscar_dispositivos(self):
        if self.ws_client.conectado:
            self.consola.registro("Dispositivo ya conectado")
            return
        self.consola.registro("Reintentando conexión al simulador...")
        self.ws_client.detener()
        self.ws_client.uri = "ws://localhost:8765"
        self.ws_client.iniciar()
    """
    ###OJO ESTA FUNCION PARA CUANDO TENGA EL ESP32 REAL!!!!
    def b_buscar_dispositivos(self):
        self.consola.registro("Buscando dispositivos...")
        self.l_estado_conexion.configure(text="◌ Buscando...", text_color="#f0c060")
        #daemon = True, significa que el hilo se detiene automáticamente cuando se cierra la App
        threading.Thread(target=self.buscar_mdns,daemon=True).start()

        #if self.ws_client.conectado:
            #self.consola.registro("DISPOSITIVO CONECTADO ✔")
        #elif

        #else:
    """
    def buscar_mdns(self):

        #se trata de un diccinario que podrá ser modificado por la clase oyente, y guarda la
        #información relacionada
        direccion_controlador = {"ip": None, "puerto" : None}

        class Oyente:
            #la funcion add_service se dispara automáticamente cuendo encuentra un dispositivo en la
            #red que coincide con el tipo de servicio que se busca
            def add_service(self, zconf, tipo, nombre):
                info =zconf.get_service_info(tipo,nombre)
                if info:
                        direccion_controlador["ip"] = info.parsed_addresses()[0]
                        direccion_controlador["puerto"] = info.port
                pass

            def remove_service(self):
                pass

            def update_service(self):
                pass

        zconf = zc.Zeroconf()
        #OJO: "_esp32._tcp.local." tiene una estructura fija
        #_esp32   → nombre del servicio (debe coincidir con el ESP32)
        #._tcp         → protocolo de transporte
        #.local.       → dominio mDNS (siempre este, con el punto final)
        #El único propósito de la variable buscador es mantener vivo el objeto 
        # en memoria durante el bucle de espera.
        buscador = zc.ServiceBrowser(zconf, "_esp32._tcp.local.", Oyente())

        """
        Como ServiceBrowser es asíncrono (trabaja en su propio hilo interno), necesitas 
        esperar a que encuentre algo. El bucle comprueba cada 100ms si ya se encontró la IP.
        Si pasan 5 segundos sin resultado, sale del bucle. zconf.close() libera el socket,
        siempre hay que cerrarlo.
        """
        tiempo_limite = 5
        tiempo_inicio = time.time()
        while direccion_controlador["ip"] is None and time.time() - tiempo_inicio < tiempo_limite:
            time.sleep(0.1)

        zconf.close()

        if direccion_controlador["ip"]:
            #Uniform Resource Identifier, un identificador único que señala la dirección de un recurso.
            #ws (WebSocket) ip y puerto
            uri = f'ws://{direccion_controlador["ip"]}:{direccion_controlador["puerto"]}' 
            self.after(0,self.conectar_dispositivo_encontrado, uri)

        else:
            self.after(0,lambda: self.consola.registro(f'No se encontró ningún dispositivo', nivel = "AVISO"))
            self.l_estado_conexion.configure(text="✕ Error",text_color="#fa8989")


    def conectar_dispositivo_encontrado(self,uri):
            #imprime/registra en la consola que se ha encontrado un dispositivo
            self.consola.registro(f'Dispositivo encontrado: {uri}')
            #para la conexión WebSocket actual
            self.ws_client.detener()
            #cambia la URI del cliente WebSocket a la del ESP32 encontrado
            self.ws_client.uri = uri
            #reconecta nuevamente con la nueva URI
            self.ws_client.iniciar()

        
    def tiempo_sesion(self):    
        tiempo = time.strftime("%H:%M:%S", time.gmtime(time.time()-self.tiempo_inicio_sesion))
        self.duracion_sesion.set(f"Duración de sesión: {tiempo}")
        self.after(1000, self.tiempo_sesion)

    def actualizar_comb_canales_calibrados(self):
        valores = [canal.e_olor_canal.get() for canal in self.cuadros_canales if canal.e_olor_canal.get()] 
        valores.append("Canal Blanco")
        self.comb_canales_calibrados.configure(values=valores)

    def bloquear_botones(self,bloquear):
        if bloquear:
            #cabecera
            self.b_buscar_dispositivos.configure(state="disabled")

            #protocolo
            self.e_num_ciclos.b_decrementar.configure(state="disabled")
            self.e_num_ciclos.e_spinbox.configure(state="disabled")
            self.e_num_ciclos.b_incrementar.configure(state="disabled")
            self.e_intervalo_ciclos.b_decrementar.configure(state="disabled")
            self.e_intervalo_ciclos.e_spinbox.configure(state="disabled")
            self.e_intervalo_ciclos.b_incrementar.configure(state="disabled")
            self.e_tiempo_exposicion.b_decrementar.configure(state="disabled")
            self.e_tiempo_exposicion.e_spinbox.configure(state="disabled")
            self.e_tiempo_exposicion.b_incrementar.configure(state="disabled")
            self.e_tiempo_desensibilizacion.b_decrementar.configure(state="disabled")
            self.e_tiempo_desensibilizacion.e_spinbox.configure(state="disabled")
            self.e_tiempo_desensibilizacion.b_incrementar.configure(state="disabled")
            self.cb_aleatorio.configure(state="disabled")
            self.cb_secuencial.configure(state="disabled")
            self.b_iniciar_protocolo.configure(state="disabled")
            self.b_reiniciar_protocolo.configure(state="disabled")
            
            #estado
            self.e_id_sesion.configure(state="disabled")
            self.e_id_paciente.configure(state="disabled")
            self.b_generar_informe.configure(state="disabled")

            #Canales
            for canal in self.cuadros_canales:
                if canal.e_olor_canal.winfo_exists():
                    canal.e_olor_canal.configure(state="disabled")
                canal.b_activar_canal.configure(state="disabled")
                canal.b_parar_canal.configure(state="disabled")
            
            #calibrado
            self.comb_canales_calibrados.configure(state="disabled")
            self.b_iniciar_calibrado_general.configure(state="disabled")
            self.b_reiniciar_calibrado_general.configure(state="disabled")
            self.b_parar_calibrado_general.configure(state="disabled")

            self.b_iniciar_calibrado_velocidad.configure(state="disabled")
            self.b_reiniciar_calibrado_velocidad.configure(state="disabled")
            self.b_parar_calibrado_velocidad.configure(state="disabled")
            self.b_iniciar_calibrado_flujo.configure(state="disabled")
            self.b_reiniciar_calibrado_flujo.configure(state="disabled")
            self.b_parar_calibrado_flujo.configure(state="disabled")
            self.b_iniciar_calibrado_concentracion.configure(state="disabled")
            self.b_reiniciar_calibrado_concentracion.configure(state="disabled")  
            self.b_parar_calibrado_concentracion.configure(state="disabled")
            self.b_iniciar_calibrado_latencia.configure(state="disabled")
            self.b_reiniciar_calibrado_latencia.configure(state="disabled")
            self.b_parar_calibrado_latencia.configure(state="disabled")


        else:

            #cabecera
            self.b_buscar_dispositivos.configure(state="normal")

            #protocolo
            self.e_num_ciclos.b_decrementar.configure(state="normal")
            self.e_num_ciclos.e_spinbox.configure(state="normal")
            self.e_num_ciclos.b_incrementar.configure(state="normal")
            self.e_intervalo_ciclos.b_decrementar.configure(state="normal")
            self.e_intervalo_ciclos.e_spinbox.configure(state="normal")
            self.e_intervalo_ciclos.b_incrementar.configure(state="normal")
            self.e_tiempo_exposicion.b_decrementar.configure(state="normal")
            self.e_tiempo_exposicion.e_spinbox.configure(state="normal")
            self.e_tiempo_exposicion.b_incrementar.configure(state="normal")
            self.e_tiempo_desensibilizacion.b_decrementar.configure(state="normal")
            self.e_tiempo_desensibilizacion.e_spinbox.configure(state="normal")
            self.e_tiempo_desensibilizacion.b_incrementar.configure(state="normal")
            self.cb_aleatorio.configure(state="normal")
            self.cb_secuencial.configure(state="normal")
            self.b_iniciar_protocolo.configure(state="normal")
            self.b_reiniciar_protocolo.configure(state="normal")
            
            #estado
            self.e_id_sesion.configure(state="normal")
            self.e_id_paciente.configure(state="normal")
            self.b_generar_informe.configure(state="normal")

            #Canales
            for canal in self.cuadros_canales:
                if canal.e_olor_canal.winfo_exists():
                    canal.e_olor_canal.configure(state="normal")
                canal.b_activar_canal.configure(state="normal")
                canal.b_parar_canal.configure(state="normal")
            
            #calibrado
            self.comb_canales_calibrados.configure(state="normal")
            self.b_iniciar_calibrado_general.configure(state="normal")
            self.b_reiniciar_calibrado_general.configure(state="normal")
            self.b_parar_calibrado_general.configure(state="normal")

            self.b_iniciar_calibrado_velocidad.configure(state="normal")
            self.b_reiniciar_calibrado_velocidad.configure(state="normal")
            self.b_parar_calibrado_velocidad.configure(state="normal")
            self.b_iniciar_calibrado_flujo.configure(state="normal")
            self.b_reiniciar_calibrado_flujo.configure(state="normal")
            self.b_parar_calibrado_flujo.configure(state="normal")
            self.b_iniciar_calibrado_concentracion.configure(state="normal")
            self.b_reiniciar_calibrado_concentracion.configure(state="normal")  
            self.b_parar_calibrado_concentracion.configure(state="normal")
            self.b_iniciar_calibrado_latencia.configure(state="normal")
            self.b_reiniciar_calibrado_latencia.configure(state="normal")
            self.b_parar_calibrado_latencia.configure(state="normal")

    def iniciar_protocolo(self):
        
        if self.protocolo_activo:
            if self.protocolo_parado:
                self.protocolo_parado = False
                self.consola.registro("Reanudando protocolo...")
                self.b_iniciar_protocolo.configure(state ="disabled")
                self.b_reiniciar_protocolo.configure(state = "disabled")

                if self.fase_actual == "Exposición":
                    canal = self.canales_protocolo[self.indice_canal_protocolo]
                    self.canal_activo = canal
                    #self.sv_canal_activo.set(self.canal_activo.e_olor_canal.get())
                    canal.activar_canal(tiempo_inicial = self.tiempo_guardado)
                    self.tiempo_guardado = None
                    self.periodo_exposicion(canal, self.segundos_restantes_protocolo)
                if self.fase_actual == "Desensibilización":
                    self.canal_activo = self.cuadros_canales[2]
                    #self.sv_canal_activo.set(self.canal_activo.e_olor_canal.get())
                    self.cuadros_canales[2].activar_canal(tiempo_inicial = self.tiempo_guardado) #CANAL DE DESENSIBILIZACIÓN
                    self.tiempo_guardado = None
                    self.periodo_desensibilizacion(self.segundos_restantes_protocolo)
                if self.fase_actual == "Intervalo":
                    self.canal_activo = None
                    self.sv_canal_activo.set("Ninguno")
                    self.sv_canal_anterior.set("Ninguno")
                    self.sv_canal_siguiente.set("Ninguno")
                    self.periodo_intervalo(self.segundos_restantes_protocolo)
            else:
                self.consola.registro("Ya hay un protocolo activo", nivel="AVISO")
            return
        
        else:
            self.protocolo_activo= True
            canales_con_olor = []

        for canal in self.cuadros_canales:
            if canal.e_olor_canal.get() != "":
                canales_con_olor.append(canal)

        #Validación
        ##ESTOS CONDICIONALES SE LOS PUEDO METER A UNA FUNCIÓN PARA QUE ASEGURE LA VALIDACIÓN DEL PROTOCOLO
        if not canales_con_olor:
            self.consola.registro("No hay ningún canal con olor introducido. Introduzca al menos un olor en uno de los canales para iniciar el protocolo", nivel="AVISO")
            self.protocolo_activo= False
            return
        self.consola.registro(f"{self.e_num_ciclos.get()}")
        if self.e_num_ciclos.get() == 0:
            self.consola.registro("El número de ciclos no está definido. Defina al menos un ciclo para iniciar el protocolo", nivel="AVISO")
            self.protocolo_activo= False
            return

        if self.e_tiempo_exposicion.get() == 0:
            self.consola.registro("El tiempo de exposición no está definido. Introduzca la duración del periodo de exposición para iniciar el protocolo", nivel="AVISO")
            self.protocolo_activo= False
            return

        if self.e_tiempo_desensibilizacion.get() == 0:
            self.consola.registro("El tiempo de desensibilización no está definido. Introduzca la duración del periodo de desensibilización para iniciar el protocolo", nivel="AVISO")
            self.protocolo_activo= False
            return

        if self.cb_aleatorio.get() == 0 and self.cb_secuencial.get() == 0:
            self.consola.registro("No se ha seleccionado el orden de los canales. Seleccione un orden para iniciar el protocolo", nivel="AVISO")
            self.protocolo_activo= False
            return

        if self.cb_aleatorio.get() == 1 and self.cb_secuencial.get() == 1:
            self.consola.registro("No se pueden seleccionar ambos órdenes de canal a la vez. Seleccione un orden para iniciar el protocolo", nivel="AVISO")
            self.protocolo_activo= False
            return

        else:
            self.ciclo_actual = 1
            self.indice_canal_protocolo = 0
            self.segundos_restantes_protocolo = 0
            self.tiempo_guardado = None
            self.fase_actual = None

            if self.cb_aleatorio.get() == 1:
                self.orden_protocolo = "aleatorio"
                self.canales_protocolo = random.sample(canales_con_olor, len(canales_con_olor))
                
            else:
                self.orden_protocolo = "secuencial"
                #hacemos una copia de la lista de olores en el protocolo y no una referencia al objeto que los guarda
                self.canales_protocolo = list(canales_con_olor)

        #ARRANQUE PROTOCOLO
        self.consola.registro(f"Iniciando protocolo {self.orden_protocolo} en {len(self.canales_protocolo)} canales: { [olor.e_olor_canal.get() for olor in self.canales_protocolo] }")
        #self.sv_estado_protocolo.set(f"Ciclo {self.ciclo_actual}/{self.e_num_ciclos.get()} - Canal {self.colores_canales[self.canal_proto_idx]} - Fase: {self.fase_actual or 'N/A'} - Tiempo restante: {self.segundos_restantes}s")
        #self.sv_cuenta_atras.set(f"{self.segundos_restantes}s")
        self.bloquear_botones(bloquear=True)
        self.iniciar_exposicion()

    def iniciar_exposicion(self):   
        if self.indice_canal_protocolo >= len(self.canales_protocolo):
            self.iniciar_intervalo()
            return
        
        canal = self.canales_protocolo[self.indice_canal_protocolo]
        self.canal_activo = canal
        self.fase_actual = "Exposición"
        self.consola.registro(f"Ciclo {self.ciclo_actual}/{self.e_num_ciclos.get()} Iniciando exposición en canal {canal.e_olor_canal.get()} durante {self.e_tiempo_exposicion.get()} segundos")
        canal.activar_canal()
        self.periodo_exposicion(canal, int(self.e_tiempo_exposicion.get()))

    def periodo_exposicion(self, canal, segundos_restantes):
        if self.protocolo_parado:
            self.segundos_restantes_protocolo = segundos_restantes
            return
        
        if segundos_restantes > 0:
            self.segundos_restantes_protocolo = segundos_restantes
            #self.sv_cuenta_atrasc.set(f"{self.segundos_restantes}s")
            self.after_activo = self.after(1000, lambda: self.periodo_exposicion(canal, segundos_restantes - 1))
        else:
            #self.sv_cuenta_atras.set("")
            canal.parar_canal()
            self.iniciar_desensibilizacion()
    
    def iniciar_desensibilizacion(self):
        self.fase_actual = "Desensibilización"
        self.consola.registro(f"Ciclo {self.ciclo_actual}/{self.e_num_ciclos.get()} Iniciando desensibilización durante {self.e_tiempo_desensibilizacion.get()} segundos")
        self.canal_activo = self.cuadros_canales[2]
        self.cuadros_canales[2].activar_canal() #CANAL DE DESENSIBILIZACIÓN
        self.periodo_desensibilizacion(int(self.e_tiempo_desensibilizacion.get()))

    def periodo_desensibilizacion(self, segundos_restantes):
        if self.protocolo_parado:
            self.segundos_restantes_protocolo = segundos_restantes
            return
        
        if segundos_restantes > 0:
            #canal.activar_canal()
            self.segundos_restantes_protocolo = segundos_restantes
            #self.sv_cuenta_atras.set(f"{self.segundos_restantes}s")
            self.after_activo = self.after(1000, lambda: self.periodo_desensibilizacion(segundos_restantes - 1))

        else:
            #self.sv_cuenta_atras.set("")
            self.cuadros_canales[2].parar_canal()
            self.indice_canal_protocolo += 1
            self.iniciar_exposicion()

    def iniciar_intervalo(self):
        self.ciclo_actual += 1
        if self.ciclo_actual > int(self.e_num_ciclos.get()):
            self.finalizar_protocolo()
            return
        self.fase_actual = "Intervalo"
        self.indice_canal_protocolo = 0
        self.sv_canal_activo.set("Ninguno")
        self.sv_canal_anterior.set("Ninguno")
        self.sv_canal_siguiente.set("Ninguno")
        self.canal_activo = None
        self.consola.registro(f"Pausa entre ciclos durante: {self.e_intervalo_ciclos.get()} segundos. Ciclo siguiente: {self.ciclo_actual}/{self.e_num_ciclos.get()}")
        self.periodo_intervalo(int(self.e_intervalo_ciclos.get()))


    def periodo_intervalo(self, segundos_restantes):
        if self.protocolo_parado:
            self.segundos_restantes_protocolo = segundos_restantes
            return
        if segundos_restantes > 0:
            self.segundos_restantes_protocolo = segundos_restantes
            #self.sv_cuenta_atras.set(f"{self.segundos_restantes}s")
            self.after_activo = self.after(1000, lambda: self.periodo_intervalo(segundos_restantes - 1))
        else:
            #self.sv_cuenta_atras.set("")
            self.iniciar_exposicion()

    def finalizar_protocolo(self):
        self.protocolo_activo = False
        self.protocolo_parado = False
        self.fase_actual = None
        #self.sv_cuenta_atras.set("")
        self.bloquear_botones(bloquear=False)
        self.consola.registro("Protocolo finalizado")
    
    def reiniciar_protocolo(self):

        for canal in self.canales_protocolo:
            canal.resetear_cronometro()

        if self.after_activo:
            self.after_cancel(self.after_activo)
            self.after_activo = None

        self.protocolo_activo = False
        self.protocolo_parado = False
        self.canal_activo = None
        self.fase_actual = None
        self.segundos_restantes_protocolo = 0
        self.tiempo_guardado = None
        self.indice_canal_protocolo = None
        self.ciclo_actual = 1
        self.canales_protocolo = []
        self.bloquear_botones(bloquear=False)
        self.sv_canal_activo.set("Ninguno")
        self.sv_canal_anterior.set("Ninguno")
        self.sv_canal_siguiente.set("Ninguno")
        self.consola.registro("Protocolo reiniciado")

    
    def parar_protocolo(self):
        self.protocolo_parado = True
        self.b_iniciar_protocolo.configure(state="normal")
        self.b_reiniciar_protocolo.configure(state="normal")
        if self.fase_actual == "Exposición" or self.fase_actual == "Desensibilización":
            self.tiempo_guardado = datetime.datetime.strptime(self.canal_activo.sv_tiempo_activo.get().split()[2], "%H:%M:%S")
        #self.consola.registro(self.canal_activo)
        for canal in self.cuadros_canales:
            canal.parar_canal()
        
        self.consola.registro("Protocolo parado")
        # Si estamos en desensibilización, guardar el restante con decimales
        #fase = getattr(self, 'fase_actual', None)
        """
        if fase == "Desensibilización" and getattr(self, 'desens_end_time', None) is not None:
            restante_f = max(0.0, self.desens_end_time - time.time())
            self.segundos_restantes_float = restante_f
            self.segundos_restantes = math.ceil(restante_f)
        if fase == "Exposición" and getattr(self, 'expos_end_time', None) is not None:
            restante_f = max(0.0, self.expos_end_time - time.time())
            self.segundos_restantes_float = restante_f
            self.segundos_restantes = math.ceil(restante_f)
        if fase == "Intervalo" and getattr(self, 'inter_end_time', None) is not None:
            restante_f = max(0.0, self.inter_end_time - time.time())
            self.segundos_restantes_float = restante_f
            self.segundos_restantes = math.ceil(restante_f)
        """

        if self.after_activo:
            #en caso de estar vigente el temporizador activo, cancelamos la llamada pendiente para evitar que se 
            # ejecute después de haber parado el protocolo
            self.after_cancel(self.after_activo)
            self.after_activo = None
        """
        if not self.protocolo_activo:
            self.consola.registro("No hay ningún protocolo activo", nivel="AVISO")
            return
        if self.protocolo_parado
            self.consola.registro()
        if protocolo_activo and 
        """

    #CALIBRADO VELOCIDAD
    def iniciar_calibrado_velocidad(self):
        if self.canal_calibrado is not None:
            if not self.calibrado_velocidad_parado and not self.calibrando_velocidad:
                self.calibrando_velocidad = True
                self.consola.registro("Iniciando calibrado de velocidad...")
                self.buffer_velocidad = collections.deque([0]*int(self.e_tiempo_calibrado.get()),maxlen=int(self.e_tiempo_calibrado.get()))
                self.tiempo_grafica_velocidad = collections.deque(range(0,int(self.e_tiempo_calibrado.get())),maxlen=int(self.e_tiempo_calibrado.get()))
                self.b_iniciar_calibrado_velocidad.configure(state="disabled")
                self.b_reiniciar_calibrado_velocidad.configure(state="disabled")
                self.canal_calibrado.activar_canal()
                self.periodo_calibrado_velocidad(self.canal_calibrado, self.e_tiempo_calibrado.get())
            else:
                self.calibrado_velocidad_parado = False
                #self.calibrando_velocidad = True
                self.consola.registro("Reanudando calibrado de velocidad...")
                self.consola.registro(f"Tiempo guardado: {self.segundos_restantes_velocidad}")
                #self.canal_calibrado.activar_canal(tiempo_inicial = self.tiempo_guardado)
                self.canal_calibrado.activar_canal()
                #self.tiempo_guardado = None
                self.periodo_calibrado_velocidad(self.canal_calibrado, self.segundos_restantes_velocidad)
                #self.sv_canal_activo.set(self.canal_activo.e_olor_canal.get())
        else:
            self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de velocidad", nivel="AVISO")
            

    def periodo_calibrado_velocidad(self, canal_calibrado, segundos_restantes):
        self.segundos_restantes_velocidad = segundos_restantes
        if segundos_restantes > 0:
            self.consola.registro(f"Calibrado de velocidad en curso... Tiempo restante: {self.segundos_restantes_velocidad}s")
            #self.sv_cuenta_atrasc.set(f"{self.segundos_restantes_velocidad}s")
            self.after_calibrado_velocidad = self.after(1000, lambda: self.periodo_calibrado_velocidad(canal_calibrado, segundos_restantes - 1))
        else:
            #self.sv_cuenta_atras.set("")
            if canal_calibrado is not None:
                canal_calibrado.parar_canal()
            else:
                self.consola.registro("No hay ningún canal seleccionado para calibrar.", nivel="AVISO")
            self.parar_calibrado_velocidad()

    def reiniciar_calibrado_velocidad(self):
        self.calibrando_velocidad = False
        self.calibrado_velocidad_parado = False
        self.segundos_restantes_velocidad = 0
        self.tiempo_guardado = None
        if self.after_calibrado_velocidad:
            self.after_cancel(self.after_calibrado_velocidad)
            self.after_calibrado_velocidad = None
        if self.canal_calibrado is not None:
            self.canal_calibrado.parar_canal()
        else:
            self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de velocidad.", nivel="AVISO")

        self.buffer_historico_calibrado_velocidad.clear()
        self.buffer_velocidad = collections.deque([0]*61, maxlen=61)
        self.tiempo_grafica_velocidad = collections.deque(list(range(0,61)), maxlen=61)

        self.ax_velocidad.clear()
        self.ax_velocidad.set_xlabel("Tiempo (s)")
        self.ax_velocidad.set_ylabel("Velocidad (m/s)")
        self.ax_velocidad.tick_params(colors = "#ffffff") 
        self.ax_velocidad.xaxis.label.set_color("#ffffff")  # Color de las etiquetas del eje x
        self.ax_velocidad.yaxis.label.set_color("#ffffff")  # Color de las etiquetas del eje y
        self.ax_velocidad.spines['bottom'].set_color('#ffffff')  # Color de la línea del eje x
        self.ax_velocidad.spines['left'].set_color('#ffffff')  # Color de
        self.ax_velocidad.spines['top'].set_color('#ffffff')  # Color de la línea superior
        self.ax_velocidad.spines['right'].set_color('#ffffff')  # Color de la línea derecha
        self.canvas_velocidad.draw()

        self.b_iniciar_calibrado_velocidad.configure(state="normal")
        self.consola.registro("Calibrado de velocidad reiniciado")
    
    def parar_calibrado_velocidad(self):
        if self.calibrando_velocidad:
            self.calibrado_velocidad_parado = True
            #self.calibrando_velocidad = False
            if self.after_calibrado_velocidad:
                self.after_cancel(self.after_calibrado_velocidad)
                self.after_calibrado_velocidad = None
            if self.canal_calibrado is not None:
                self.canal_calibrado.parar_canal()
            else:
                self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de velocidad.", nivel="AVISO")
            self.b_iniciar_calibrado_velocidad.configure(state="normal")
            self.b_reiniciar_calibrado_velocidad.configure(state="normal")

            if self.segundos_restantes_velocidad > 0:
                self.consola.registro(f"Calibrado de velocidad parado con {self.segundos_restantes_velocidad} segundos restantes")
        
            else:
                if self.canal_calibrado is not None:
                    num_canal = self.canal_calibrado.num_canal
                    if num_canal not in self.metricas_calibracion:
                        self.metricas_calibracion[num_canal] = {}
                        self.metricas_calibracion[num_canal]["olor"] = self.canal_calibrado.e_olor_canal.get() or self.canal_calibrado.color_canal
                    self.metricas_calibracion[num_canal]["velocidad"] = list(self.buffer_historico_calibrado_velocidad)
                    self.calibrando_velocidad = False
                    self.calibrado_velocidad_parado = False
                    self.consola.registro("Calibrado de velocidad finalizado")
                else:
                    self.consola.registro("No se pudieron guardar las métricas de calibrado: No hay ningún canal seleccionado para calibrar.", nivel="AVISO")
        else:
            self.consola.registro("No hay ningún calibrado de velocidad activo. Seleccione un canal y pulse iniciar para comenzar el calibrado de velocidad.", nivel="AVISO")
            

    #CALIBRADO FLUJO
    def iniciar_calibrado_flujo(self):
        if self.canal_calibrado is not None:
            if not self.calibrado_flujo_parado and not self.calibrando_flujo:
                self.calibrando_flujo = True
                self.consola.registro("Iniciando calibrado de flujo...")
                self.buffer_flujo = collections.deque([0]*int(self.e_tiempo_calibrado.get()),maxlen=int(self.e_tiempo_calibrado.get()))
                self.tiempo_grafica_flujo = collections.deque(range(int(self.e_tiempo_calibrado.get())),maxlen=int(self.e_tiempo_calibrado.get()))
                self.b_iniciar_calibrado_flujo.configure(state="disabled")
                self.b_reiniciar_calibrado_flujo.configure(state="disabled")
                self.canal_calibrado.activar_canal()
                self.periodo_calibrado_flujo(self.canal_calibrado, self.e_tiempo_calibrado.get())
            else:
                self.calibrado_flujo_parado = False
                #self.calibrando_flujo = True
                self.consola.registro("Reanudando calibrado de flujo...")
                self.consola.registro(f"Tiempo guardado: {self.segundos_restantes_flujo}")
                #self.canal_calibrado.activar_canal(tiempo_inicial = self.tiempo_guardado)
                self.canal_calibrado.activar_canal()
                #self.tiempo_guardado = None
                self.periodo_calibrado_flujo(self.canal_calibrado, self.segundos_restantes_flujo)
                #self.sv_canal_activo.set(self.canal_activo.e_olor_canal.get())
        else:
            self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de flujo", nivel="AVISO")
            
        
    def periodo_calibrado_flujo(self, canal_calibrado, segundos_restantes):
        self.segundos_restantes_flujo = segundos_restantes
        if segundos_restantes > 0:
            self.consola.registro(f"Calibrado de flujo en curso... Tiempo restante: {self.segundos_restantes_flujo}s")
            #self.sv_cuenta_atrasc.set(f"{self.segundos_restantes_flujo}s")
            self.after_calibrado_flujo = self.after(1000, lambda: self.periodo_calibrado_flujo(canal_calibrado, segundos_restantes - 1))
        else:
            #self.sv_cuenta_atras.set("")
            if canal_calibrado is not None:
                canal_calibrado.parar_canal()
            else:
                self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de flujo.", nivel="AVISO")
            self.parar_calibrado_flujo()

    def reiniciar_calibrado_flujo(self):
        self.calibrando_flujo = False
        self.calibrado_flujo_parado = False
        self.segundos_restantes_flujo = 0
        self.tiempo_guardado = None
        if self.after_calibrado_flujo:
            self.after_cancel(self.after_calibrado_flujo)
            self.after_calibrado_flujo = None
        if self.canal_calibrado is not None:
            self.canal_calibrado.parar_canal()
        else:
            self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de flujo.", nivel="AVISO")

        self.buffer_historico_calibrado_flujo.clear()
        self.buffer_flujo = collections.deque([0]*61, maxlen=61)
        self.tiempo_grafica_flujo = collections.deque(list(range(0,61)), maxlen=61)

        self.ax_flujo.clear()
        self.ax_flujo.set_xlabel("Tiempo (s)")
        self.ax_flujo.set_ylabel("Flujo (l/min)")
        self.ax_flujo.set_title("Flujo")
        self.canvas_flujo.draw()

        self.b_iniciar_calibrado_flujo.configure(state="normal")
        self.consola.registro("Calibrado de flujo reiniciado")
    
    def parar_calibrado_flujo(self):
        if self.calibrando_flujo:
            self.calibrado_flujo_parado = True
            #self.calibrando_flujo = False
            if self.after_calibrado_flujo:
                self.after_cancel(self.after_calibrado_flujo)
                self.after_calibrado_flujo = None
            if self.canal_calibrado is not None:
                self.canal_calibrado.parar_canal()
            else:
                self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de flujo.", nivel="AVISO")
            self.b_iniciar_calibrado_flujo.configure(state="normal")
            self.b_reiniciar_calibrado_flujo.configure(state="normal")

            if self.segundos_restantes_flujo > 0:
                #self.tiempo_guardado = datetime.datetime.strptime(int(self.e_tiempo_calibrado.get()) - self.segundos_restantes, "%H:%M:%S")
                #self.consola.registro(f"{self.tiempo_guardado}")
                self.consola.registro(f"Calibrado de flujo parado con {self.segundos_restantes_flujo} segundos restantes")
        
            else:
                if self.canal_calibrado is not None:
                    num_canal = self.canal_calibrado.num_canal
                    if num_canal not in self.metricas_calibracion:
                        self.metricas_calibracion[num_canal] = {}
                        self.metricas_calibracion[num_canal]["olor"] = self.canal_calibrado.e_olor_canal.get() or self.canal_calibrado.color_canal
                    self.metricas_calibracion[num_canal]["flujo"] = list(self.buffer_historico_calibrado_flujo)
                    self.calibrando_flujo = False
                    self.calibrado_flujo_parado = False
                    self.consola.registro("Calibrado de flujo finalizado")
                else:
                    self.consola.registro("No se pudieron guardar métricas: No hay ningún canal seleccionado para calibrar. ", nivel="AVISO")
        else:
            self.consola.registro("No hay ningún calibrado activo. Seleccione un canal y pulse iniciar para comenzar el calibrado de flujo.", nivel="AVISO")
    
    #CALIBRADO CONCENTRACIÓN
    def iniciar_calibrado_concentracion(self):
        if self.canal_calibrado is not None:
            if not self.calibrado_concentracion_parado and not self.calibrando_concentracion:
                self.calibrando_concentracion = True
                self.consola.registro("Iniciando calibrado de concentración...")
                self.buffer_concentracion = collections.deque([0]*int(self.e_tiempo_calibrado.get()),maxlen=int(self.e_tiempo_calibrado.get()))
                self.tiempo_grafica_concentracion = collections.deque(range(0,int(self.e_tiempo_calibrado.get())),maxlen=int(self.e_tiempo_calibrado.get()))
                self.b_iniciar_calibrado_concentracion.configure(state="disabled")
                self.b_reiniciar_calibrado_concentracion.configure(state="disabled")
                self.canal_calibrado.activar_canal()
                self.periodo_calibrado_concentracion(self.canal_calibrado, self.e_tiempo_calibrado.get())
            else:
                self.calibrado_concentracion_parado = False
                #self.calibrando_concentracion = True
                self.consola.registro("Reanudando calibrado de concentración...")
                self.consola.registro(f"Tiempo guardado: {self.segundos_restantes_concentracion}")
                #self.canal_calibrado.activar_canal(tiempo_inicial = self.tiempo_guardado)
                self.canal_calibrado.activar_canal()
                #self.tiempo_guardado = None
                self.periodo_calibrado_concentracion(self.canal_calibrado, self.segundos_restantes_concentracion)
                #self.sv_canal_activo.set(self.canal_activo.e_olor_canal.get())
        else:
            self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de concentración", nivel="AVISO")
            
        
    def periodo_calibrado_concentracion(self, canal_calibrado, segundos_restantes):
        self.segundos_restantes_concentracion = segundos_restantes
        if segundos_restantes > 0:
            self.consola.registro(f"Calibrado de concentración en curso... Tiempo restante: {self.segundos_restantes_concentracion}s")
            #self.sv_cuenta_atrasc.set(f"{self.segundos_restantes}s")
            self.after_calibrado_concentracion = self.after(1000, lambda: self.periodo_calibrado_concentracion(canal_calibrado, segundos_restantes - 1))
        else:
            #self.sv_cuenta_atras.set("")
            if canal_calibrado is not None:
                canal_calibrado.parar_canal()
            else:
                self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de concentración.", nivel="AVISO")
            self.parar_calibrado_concentracion()

    def reiniciar_calibrado_concentracion(self):
        self.calibrando_concentracion = False
        self.calibrado_concentracion_parado = False
        self.segundos_restantes_concentracion = 0
        self.tiempo_guardado = None
        if self.after_calibrado_concentracion:
            self.after_cancel(self.after_calibrado_concentracion)
            self.after_calibrado_concentracion = None
        if self.canal_calibrado is not None:
            self.canal_calibrado.parar_canal()
        else:
            self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de concentración.", nivel="AVISO")

        self.buffer_historico_calibrado_concentracion.clear()
        self.buffer_concentracion = collections.deque([0]*61, maxlen=61)
        self.tiempo_graficas_concentracion = collections.deque(list(range(0,61)), maxlen=61)

        self.ax_concentracion.clear()
        self.ax_concentracion.set_xlabel("Tiempo (s)")
        self.ax_concentracion.set_ylabel("Concentración (ppm)")
        self.ax_concentracion.set_title("Concentración")
        self.canvas_concentracion.draw()

        self.b_iniciar_calibrado_concentracion.configure(state="normal")
        self.consola.registro("Calibrado de concentración reiniciado")

    def parar_calibrado_concentracion(self):
        if self.calibrando_concentracion:
            self.calibrado_concentracion_parado = True
            #self.calibrando_concentracion = False
            if self.after_calibrado_concentracion:
                self.after_cancel(self.after_calibrado_concentracion)
                self.after_calibrado_concentracion = None
            if self.canal_calibrado is not None:
                self.canal_calibrado.parar_canal()
            else:
                self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de concentración.", nivel="AVISO")
            self.b_iniciar_calibrado_concentracion.configure(state="normal")
            self.b_reiniciar_calibrado_concentracion.configure(state="normal")

            if self.segundos_restantes_concentracion > 0:
                #self.tiempo_guardado = datetime.datetime.strptime(int(self.e_tiempo_calibrado.get()) - self.segundos_restantes, "%H:%M:%S")
                #self.consola.registro(f"{self.tiempo_guardado}")
                self.consola.registro(f"Calibrado de concentración parado con {self.segundos_restantes_concentracion} segundos restantes")
        
            else:
                if self.canal_calibrado is not None:
                    num_canal = self.canal_calibrado.num_canal
                    if num_canal not in self.metricas_calibracion:
                        self.metricas_calibracion[num_canal] = {}
                        self.metricas_calibracion[num_canal]["olor"] = self.canal_calibrado.e_olor_canal.get() or self.canal_calibrado.color_canal
                    self.metricas_calibracion[num_canal]["concentracion"] = list(self.buffer_historico_calibrado_concentracion)
                    self.calibrando_concentracion = False
                    self.calibrado_concentracion_parado = False
                    self.consola.registro("Calibrado de concentración finalizado")
                else:
                    self.consola.registro("No se pudieron guardar métricas: No hay ningún canal seleccionado para calibrar.", nivel="AVISO")
        else:
            self.consola.registro("No hay ningún calibrado activo. Seleccione un canal y pulse iniciar para comenzar el calibrado de concentración.", nivel="AVISO")

    #CALIBRADO LATENCIA
    def iniciar_calibrado_latencia(self):
        if self.canal_calibrado is not None:
            if not self.calibrado_latencia_parado and not self.calibrando_latencia:
                self.calibrando_latencia = True
                self.consola.registro("Iniciando calibrado de latencia...")
                self.buffer_latencia = collections.deque([0]*int(self.e_tiempo_calibrado.get()),maxlen=int(self.e_tiempo_calibrado.get()))
                self.tiempo_grafica_latencia = collections.deque(range(0,int(self.e_tiempo_calibrado.get())),maxlen=int(self.e_tiempo_calibrado.get()))
                self.b_iniciar_calibrado_latencia.configure(state="disabled")
                self.b_reiniciar_calibrado_latencia.configure(state="disabled")
                self.canal_calibrado.activar_canal()
                self.periodo_calibrado_latencia(self.canal_calibrado, self.e_tiempo_calibrado.get())
            else:
                self.calibrado_latencia_parado = False
                #self.calibrando_latencia = True
                self.consola.registro("Reanudando calibrado de latencia...")
                self.consola.registro(f"Tiempo guardado: {self.segundos_restantes_latencia}")
                #self.canal_calibrado.activar_canal(tiempo_inicial = self.tiempo_guardado)
                self.canal_calibrado.activar_canal()
                #self.tiempo_guardado = None
                self.periodo_calibrado_latencia(self.canal_calibrado, self.segundos_restantes_latencia)
                #self.sv_canal_activo.set(self.canal_activo.e_olor_canal.get())
        else:
            self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de latencia", nivel="AVISO")
            

    def periodo_calibrado_latencia(self, canal_calibrado, segundos_restantes):
        self.segundos_restantes_latencia = segundos_restantes
        if segundos_restantes > 0:
            self.consola.registro(f"Calibrado de latencia en curso... Tiempo restante: {self.segundos_restantes_latencia}s")
            #self.sv_cuenta_atrasc.set(f"{self.segundos_restantes_latencia}s")
            self.after_calibrado_latencia = self.after(1000, lambda: self.periodo_calibrado_latencia(canal_calibrado, segundos_restantes - 1))
        else:
            #self.sv_cuenta_atras.set("")
            if canal_calibrado is not None:
                canal_calibrado.parar_canal()
            else:
                self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de latencia.", nivel="AVISO")
            self.parar_calibrado_latencia()

    def reiniciar_calibrado_latencia(self):
        self.calibrando_latencia = False
        self.calibrado_latencia_parado = False
        self.segundos_restantes_latencia = 0
        self.tiempo_guardado = None
        if self.after_calibrado_latencia:
            self.after_cancel(self.after_calibrado_latencia)
            self.after_calibrado_latencia = None
        if self.canal_calibrado is not None:
            self.canal_calibrado.parar_canal()
        else:
            self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de latencia.", nivel="AVISO")

        self.buffer_historico_calibrado_latencia.clear()
        self.buffer_latencia = collections.deque([0]*61, maxlen=61)
        self.tiempo_grafica_latencia = collections.deque(list(range(0,61)), maxlen=61)

        self.ax_latencia.clear()
        self.ax_latencia.set_xlabel("Tiempo (s)")
        self.ax_latencia.set_ylabel("Latencia (ms)")
        self.ax_latencia.set_title("Latencia")
        self.canvas_latencia.draw()

        self.b_iniciar_calibrado_latencia.configure(state="normal")
        self.consola.registro("Calibrado de latencia reiniciado")

    def parar_calibrado_latencia(self):
        if self.calibrando_latencia:
            self.calibrado_latencia_parado = True
            #self.calibrando_latencia = False
            if self.after_calibrado_latencia:
                self.after_cancel(self.after_calibrado_latencia)
                self.after_calibrado_latencia = None
            if self.canal_calibrado is not None:
                self.canal_calibrado.parar_canal()
            else:
                self.consola.registro("No hay ningún canal seleccionado para calibrar. Seleccione un canal para iniciar el calibrado de latencia.", nivel="AVISO")
            self.b_iniciar_calibrado_latencia.configure(state="normal")
            self.b_reiniciar_calibrado_latencia.configure(state="normal")

            if self.segundos_restantes_latencia > 0:
                #self.tiempo_guardado = datetime.datetime.strptime(int(self.e_tiempo_calibrado.get()) - self.segundos_restantes, "%H:%M:%S")
                #self.consola.registro(f"{self.tiempo_guardado}")
                self.consola.registro(f"Calibrado de latencia parado con {self.segundos_restantes_latencia} segundos restantes")
        
            else:
                num_canal = self.canal_calibrado.num_canal
                if num_canal not in self.metricas_calibracion:
                    self.metricas_calibracion[num_canal] = {}
                    self.metricas_calibracion[num_canal]["olor"] = self.canal_calibrado.e_olor_canal.get() or self.canal_calibrado.color_canal
                self.metricas_calibracion[num_canal]["latencia"] = list(self.buffer_historico_calibrado_latencia)
                self.calibrando_latencia = False
                self.calibrado_latencia_parado = False
                self.consola.registro("Calibrado de latencia finalizado")
        else:
            self.consola.registro("No hay ningún calibrado activo. Seleccione un canal y pulse iniciar para comenzar el calibrado de latencia.", nivel="AVISO")


    def iniciar_calibrado_general(self):
        self.iniciar_calibrado_velocidad()
        self.iniciar_calibrado_flujo()
        self.iniciar_calibrado_concentracion()
        self.iniciar_calibrado_latencia()
    def reiniciar_calibrado_general(self):
        self.reiniciar_calibrado_velocidad()
        self.reiniciar_calibrado_flujo()
        self.reiniciar_calibrado_concentracion()
        self.reiniciar_calibrado_latencia()
    def parar_calibrado_general(self):
        self.parar_calibrado_velocidad()
        self.parar_calibrado_flujo()
        self.parar_calibrado_concentracion()
        self.parar_calibrado_latencia()
    
    def media(self, lista):
        # Acepta tanto listas/iterables como valores numéricos.
        # Si `lista` es un número, lo devolvemos tal cual.
        try:
            if lista is None:
                return 0
            if isinstance(lista, (int, float)):
                return lista
            # Evitar calcular la media de strings
            if isinstance(lista, str):
                return 0
            # Intentar iterar y calcular la media
            seq = list(lista)
            return statistics.mean(seq) if seq else 0
        
        except Exception as e:
            self.consola.registro(f"Error al calcular la media: {e}", nivel="ERROR")
            return None

    def protocolo_definido(self,protocolo):
        if protocolo != '------':
            self.consola.registro(f'Protolo definido seleccionado: {protocolo}')
        else:
            self.consola.registro(f'Prorocolo definido no seleccionado')
    
    def seleccionar_calibrado(self,nombre_canal):
        if not self.calibrado_activo():
            if nombre_canal != 'Ninguno':
                self.consola.registro(f'Canal calibrado seleccionado: {nombre_canal}')
                for canal in self.cuadros_canales:
                    if canal.e_olor_canal.get() == nombre_canal or canal.l_color_canal.cget("text") == nombre_canal:
                        self.canal_calibrado = canal
            else:
                self.consola.registro(f'Canal calibrado no seleccionado')

        else:   
            self.comb_canales_calibrados.configure(state = "disabled")
        
                
    def calibrado_activo(self):
        return (self.calibrando_velocidad or self.calibrando_flujo or
                self.calibrando_concentracion or self.calibrando_latencia)
        
    def actualizar_graficas(self):
        #self.buffer_velocidad.append(nuevo_valor_velocidad)
        #self.buffer_flujo.append(nuevo_valor_flujo)
        #self.buffer_concentración.append(nuevo_valor_concentración)
        #self.buffer_latencia.append(nuevo_valor_latencia)
        if self.calibrando_velocidad and not self.calibrado_velocidad_parado:
            self.ax_velocidad.clear()
            self.ax_velocidad.plot(self.tiempo_grafica_velocidad, self.buffer_velocidad)
            self.ax_velocidad.set_xlabel("Tiempo (s)")
            self.ax_velocidad.set_ylabel("Velocidad (m/s)")
            self.canvas_velocidad.draw()
        if self.calibrando_flujo and not self.calibrado_flujo_parado:
            self.ax_flujo.clear()
            self.ax_flujo.plot(self.tiempo_grafica_flujo, self.buffer_flujo)
            self.ax_flujo.set_xlabel("Tiempo (s)")
            self.ax_flujo.set_ylabel("Flujo (ml/min)")
            self.canvas_flujo.draw()
        if self.calibrando_concentracion and not self.calibrado_concentracion_parado:
            self.ax_concentracion.clear()
            self.ax_latencia.clear()
            self.ax_concentracion.plot(self.tiempo_grafica_concentracion, self.buffer_concentracion)
            self.ax_concentracion.set_xlabel("Tiempo (s)")
            self.ax_concentracion.set_ylabel("Concentración (µg/m\u00B3)")
            self.canvas_concentracion.draw()
        if self.calibrando_latencia and not self.calibrado_latencia_parado:
            self.ax_latencia.clear()
            self.ax_latencia.plot(self.tiempo_grafica_latencia, self.buffer_latencia)
            self.ax_latencia.set_xlabel("Tiempo (s)")
            self.ax_latencia.set_ylabel("Latencia (ms)")
            self.ax_latencia.set_title("Latencia")
            self.canvas_latencia.draw()
            
        self.after_actualizar_graficas = self.after(1000, self.actualizar_graficas)
    


    def consultar_colores(self):
        for canal in self.cuadros_canales:
            self.olores.append(canal.e_olor_canal.get())

    #NO FUNCIONA LA FUNCIÓN
    #HAY QUE VOLVER A IMPLEMENTAR EL MÉTODO
   
    def actualizar_canales(self,num_canal,accion):     
            if accion== "activar":
                if self.canal_activo != None and self.canal_activo.num_canal != num_canal:
                    self.self.canal_activo.parar_canal()
                self.canal_activo = self.cuadros_canales[num_canal]
                self.sv_canal_activo.set(f"Canal {self.canal_activo.e_olor_canal.get()}" if self.canal_activo.e_olor_canal.get() else "Canal Blanco")

                if self.protocolo_activo:

                    if self.sv_canal_activo.get() == "Canal Blanco":
                        self.sv_canal_anterior.set(f"Canal {self.canales_protocolo[self.indice_canal_protocolo].e_olor_canal.get()}" if self.indice_canal_protocolo >=0 else "Ninguno")
                        self.sv_canal_siguiente.set(f"Canal {self.canales_protocolo[self.indice_canal_protocolo + 1].e_olor_canal.get()}" if self.indice_canal_protocolo+1 < len(self.canales_protocolo)  else "Ninguno")
                    else:
                        self.sv_canal_anterior.set(f"Canal Blanco" if self.indice_canal_protocolo - 1 >= 0 else "Ninguno")
                        self.sv_canal_siguiente.set(f"Canal Blanco" if self.indice_canal_protocolo + 1 <= len(self.canales_protocolo)  else "Ninguno")      

                self.ws_client.enviar({"cmd": "activar", "canal": num_canal, "velocidad_%": 100})
            
            if accion== "parar":
                if self.canal_activo.num_canal == num_canal:
                    self.canal_activo = None
                    self.sv_canal_activo.set("Ninguno")
                    #podrían establecerse como valor "Ninguno" a los canales anterior y siguiente.

                self.ws_client.enviar({"cmd": "parar", "canal": num_canal})
        

            #self.after(1000, self.actualizar_canales)
            

#FUNCIONES NECESARIAS PARA LA CONCEXIÓN CON ESP32
    def _on_datos_ws(self, datos: dict):
# Llamado desde el hilo WS → usar after() para tocar la UI
        canal = datos.get("canal", -1)
        # Actualizar buffers de la gráfica correspondiente
        if self.canal_activo is not None and canal == self.canal_activo.num_canal:

            self.buffer_historico_flujo.append(datos["flujo"])
            self.buffer_historico_velocidad.append(datos["velocidad_motor"])
            self.buffer_historico_concentracion.append(datos["concentracion"])
            self.buffer_historico_latencia.append(datos["latencia"])
            self.buffer_historico_timestamps.append(datos["timestamp"])

            try:
                olor = self.canal_activo.e_olor_canal.get()
            except Exception:
                olor = ""
            self.buffer_historico_olores.append(olor)
            datos_con_hist = dict(datos)
            datos_con_hist["olor"] = olor
            self.historial_sesion.append(datos_con_hist)
        
        if self.canal_calibrado is not None and canal == self.canal_calibrado.num_canal:
            if self.calibrando_flujo:
                self.buffer_flujo.append(datos["flujo"])
                self.buffer_historico_calibrado_flujo.append(datos["flujo"])
            if self.calibrando_concentracion:
                self.buffer_concentracion.append(datos["concentracion"])
                self.buffer_historico_calibrado_concentracion.append(datos["concentracion"])
            if self.calibrando_latencia:
                self.buffer_latencia.append(datos["latencia"])
                self.buffer_historico_calibrado_latencia.append(datos["latencia"])
            if self.calibrando_velocidad:
                self.buffer_velocidad.append(datos["velocidad_motor"])
                self.buffer_historico_calibrado_velocidad.append(datos["velocidad_motor"])

        # Actualizar labels de Estado
        self.after(0, self._actualizar_labels_estado, datos)


    def _actualizar_labels_estado(self, datos):
        if datos.get("canal") == self.canal_activo:
            self.sv_flujo_aire_canal.set(f"{datos['flujo']:.1f} ml/min")
            self.sv_concentracion_canal.set(f"{datos['concentracion']:.1f} µg/m\u00B3")
            self.sv_latencia_canal.set(f"{datos['latencia']} ms")

    def _on_estado_ws(self, estado: str):
        # Actualizar el indicador de conexión en la cabecera
        textos = {
            "conectando":    ("◌ Conectando…",  "#f0c060"),
            "conectado":     ("● Conectado",    "#7DEB7D"),
            "desconectado":  ("○ Desconectado", "#fa8989"),
            "error":         ("✕ Error",        "#fa8989"),
        }
        texto, color = textos.get(estado, ("○ Desconectado", "#fa8989"))
        self.after(0, lambda: self.l_estado_conexion.configure(text=texto, text_color=color))

if __name__ == "__main__":  
    app = App()
    app.mainloop()




